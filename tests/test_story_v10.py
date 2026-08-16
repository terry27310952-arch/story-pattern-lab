from __future__ import annotations

import sys
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "apps" / "streamlit"))

import mode_exporter_v5  # noqa: E402
import story_article_cleaner  # noqa: E402
import story_content_pipeline_v5  # noqa: E402
import story_graph_engine  # noqa: E402
import story_renderer_v5  # noqa: E402
import story_source_engine_v5  # noqa: E402


TRANSFORMATION = {
    "id": "generic-transform-1",
    "source": "Independent Business Desk",
    "source_type": "rss",
    "region": "Global",
    "category": "Infrastructure",
    "title": "NeoGrid Holdings、GPUホスティングからAI計算基盤へ事業を拡大",
    "url": "https://example.com/neogrid",
    "tags": "INFRASTRUCTURE, AI, COMPUTE",
    "trader_score": 40,
    "risk_score": 10,
    "excerpt": "NeoGrid Holdingsが既存設備をAI計算基盤へ転用する。",
    "material": (
        "NeoGrid HoldingsはこれまでGPUホスティングを主力事業としてきた。 "
        "同社は既存の電力設備をAI計算基盤へ転用し、大手クラウド企業と10年間の利用契約を締結した。 "
        "契約対象は420MWで、契約価値は約1200億円。 "
        "2028年に第1期を稼働し、2030年までに全設備の稼働を計画している。 "
        "AI向け計算需要の増加を受け、既存設備の収益源を広げる。"
    ),
}

POLICY = {
    "id": "generic-policy-1",
    "source": "Public Policy Desk",
    "source_type": "rss",
    "region": "Japan",
    "category": "Policy",
    "title": "新しいデジタル資産保管ルール、2027年から段階適用へ",
    "url": "https://example.com/policy",
    "tags": "POLICY, DIGITAL ASSET",
    "trader_score": 42,
    "risk_score": 12,
    "excerpt": "金融当局がデジタル資産保管ルールの段階適用を公表した。",
    "material": (
        "金融当局はデジタル資産の保管事業者に対する新しい規制枠組みを公表した。 "
        "新ルールは2027年4月から段階的に施行する予定だ。 "
        "対象事業者には顧客資産の分別管理と監査報告が求められる。 "
        "既存事業者には12か月の移行期間を設ける。"
    ),
}

SECOND_PUBLISHER_SAME_EVENT = {
    **TRANSFORMATION,
    "id": "generic-transform-2",
    "source": "Another Editorial Desk",
    "url": "https://example.net/neogrid-deal",
    "title": "NeoGrid Holdings、420MWのAI計算基盤契約を締結",
    "material": (
        "NeoGrid Holdingsは大手クラウド企業とAI計算基盤の利用契約を締結した。 "
        "対象は420MWで、契約価値は約1200億円。2028年から段階稼働する。"
    ),
}

SAYLOR_STYLE = {
    "id": "english-store-value",
    "source": "Global Desk",
    "source_type": "rss",
    "region": "Global",
    "category": "Macro",
    "title": "Bitcoin Is a ‘Deep Freeze’ for Money. What Does That Actually Mean?",
    "url": "https://example.com/store-value",
    "tags": "BTC, MACRO",
    "trader_score": 87,
    "risk_score": 18,
    "excerpt": "Michael Saylor compares Bitcoin with a long-term store of value.",
    "material": (
        "MicroStrategy founder Michael Saylor explains Bitcoin as a long-term store of value. "
        "Gold has historically served as a store of value, while Bitcoin follows programmed scarcity. "
        "Money is energy. Bitcoin is digital monetary energy. https://t.co/example — Michael Saylor (@saylor) August 15, 2026 "
        "BTC currently trades near $63,000. "
        "Still, the analogy explains the investment thesis clearly: Bitcoin’s main pitch is preserving purchasing power across time wit"
    ),
}

RIOT_MIXED_LANGUAGE = {
    "id": "jp-riot",
    "source": "Japanese Finance Desk",
    "source_type": "rss",
    "region": "Japan",
    "category": "Infrastructure",
    "title": "BTCマイナー企業Riot、アンソロピックとデータセンター契約か",
    "url": "https://example.jp/riot",
    "tags": "BTC, AI",
    "trader_score": 45,
    "risk_score": 12,
    "excerpt": "RiotがAIインフラへ事業を拡大する。",
    "material": (
        "Riotはビットコイン採掘を主力としてきた。 "
        "Riotはアンソロピック向けのデータセンター契約を締結し、AIインフラへ事業を拡大する。 "
        "契約対象は191MWで、20年間の長期契約となる。 "
        "2027年に第1期を稼働し、2028年に全容量の稼働を計画している。 "
        "アンソロピックはAI向け計算能力を拡大している。"
    ),
}


class StoryV10Test(unittest.TestCase):
    def test_generic_source_engine_has_no_named_company_registry_or_archetype_classifier(self) -> None:
        source = (ROOT / "apps" / "streamlit" / "story_source_engine_v5.py").read_text(encoding="utf-8")
        for forbidden in ["Riot Platforms", "Anthropic", "BlackRock", "Coinspeaker", "KNOWN_ENTITIES", "classify_archetype"]:
            self.assertNotIn(forbidden, source)
        row = story_source_engine_v5.annotate_resource(TRANSFORMATION)
        self.assertEqual(row["story_archetype_hint"], "dynamic")
        self.assertGreater(row["story_score"], 0)
        self.assertIn("NeoGrid Holdings", row["story_entities"])

    def test_multilingual_entity_extraction_rejects_title_phrases(self) -> None:
        english = story_source_engine_v5.annotate_resource(story_article_cleaner.clean_story_resource(SAYLOR_STYLE))
        names = set(english["story_entities"])
        self.assertIn("Michael Saylor", names)
        self.assertFalse({"Money", "Deep Freeze", "Actually Mean", "Bitcoin Is"} & names)

        japanese = story_source_engine_v5.annotate_resource(RIOT_MIXED_LANGUAGE)
        self.assertIn("Riot", japanese["story_entities"])
        self.assertIn("アンソロピック", japanese["story_entities"])
        self.assertNotIn("マイナー", japanese["story_entities"])

    def test_generic_event_cluster_uses_event_evidence_not_publisher(self) -> None:
        a, b = story_source_engine_v5.annotate_resources([TRANSFORMATION, SECOND_PUBLISHER_SAME_EVENT])
        self.assertGreaterEqual(story_source_engine_v5.event_similarity(a, b), 0.47)
        clusters = story_source_engine_v5.cluster_story_candidates([TRANSFORMATION, SECOND_PUBLISHER_SAME_EVENT])
        self.assertEqual(len(clusters), 1)

    def test_cleaner_removes_social_embed_and_incomplete_tail(self) -> None:
        cleaned = story_article_cleaner.clean_story_resource(SAYLOR_STYLE)
        material = cleaned["material"]
        diag = cleaned["story_cleaning"]
        self.assertNotIn("t.co/", material)
        self.assertNotIn("@saylor", material)
        self.assertNotIn("across time wit", material)
        self.assertGreaterEqual(diag["social_embed_removed"], 1)
        self.assertTrue(diag["incomplete_tail_dropped"])

    def test_graph_extracts_typed_values_and_semantic_relations(self) -> None:
        hero = {"resource_ids": [TRANSFORMATION["id"]], "entities": ["NeoGrid Holdings"]}
        graph = story_graph_engine.extract_fact_graph(hero, [TRANSFORMATION])
        relations = {f["relation"] for f in graph["facts"]}
        self.assertIn("before_state", relations)
        self.assertTrue({"change", "deal"} & relations)
        self.assertIn("future", relations)
        blob = str(graph)
        self.assertIn("420MW", blob)
        self.assertIn("1200億円", blob)
        self.assertIn("2028", blob)
        self.assertIn("2030", blob)

    def test_symbol_money_is_scale_but_year_alone_is_not(self) -> None:
        row = story_article_cleaner.clean_story_resource(SAYLOR_STYLE)
        hero = {"resource_ids": [row["id"]], "entities": ["Michael Saylor"], "hero_resource": row}
        graph = story_graph_engine.extract_fact_graph(hero, [row])
        price_facts = [
            f for f in graph["facts"]
            if any(d.get("kind") == "money" and d.get("raw") == "$63,000" for d in f.get("value_details") or [])
        ]
        self.assertTrue(price_facts)
        plan = story_graph_engine.build_story_plan(hero, graph, 5)
        for item in plan["cards"]:
            if item["role"] == "scale":
                fact = next(f for f in graph["facts"] if f["id"] == item["fact_ids"][0])
                self.assertTrue(any(d.get("kind") in {"money", "capacity", "percent", "quantity", "price", "valuation"} for d in fact.get("value_details") or []))
                self.assertFalse(all(d.get("kind") == "year" for d in fact.get("value_details") or []))

    def test_investment_thesis_is_not_a_deal(self) -> None:
        self.assertNotIn("deal", story_graph_engine._relations("Still, the analogy explains the investment thesis clearly."))

    def test_purchasing_power_is_not_electrical_infrastructure(self) -> None:
        fact = {"sentence": "Bitcoin preserves purchasing power across time.", "value_details": []}
        self.assertEqual(story_graph_engine.infer_scene_type("evidence", [fact]), "asset_store_of_value")
        self.assertNotEqual(story_graph_engine.infer_scene_type("evidence", [fact]), "industrial_infrastructure")

    def test_single_historical_phrase_does_not_force_historical_parallel(self) -> None:
        row = story_article_cleaner.clean_story_resource(SAYLOR_STYLE)
        annotated = story_source_engine_v5.annotate_resource(row)
        hero = {"resource_ids": [row["id"]], "entities": annotated["story_entities"], "hero_resource": row}
        graph = story_graph_engine.extract_fact_graph(hero, [row])
        plan = story_graph_engine.build_story_plan(hero, graph, 5)
        self.assertNotEqual(plan["archetype_tag"], "historical_parallel")

    def test_story_plan_is_fact_first_and_archetype_is_only_a_tag(self) -> None:
        annotated = story_source_engine_v5.annotate_resource(TRANSFORMATION)
        hero = {"resource_ids": [TRANSFORMATION["id"]], "entities": annotated["story_entities"], "hero_resource": TRANSFORMATION, "headline_ja": TRANSFORMATION["title"]}
        graph = story_graph_engine.extract_fact_graph(hero, [TRANSFORMATION])
        plan = story_graph_engine.build_story_plan(hero, graph, 6)
        self.assertEqual(plan["planning_policy"], "facts first -> typed values/relations -> dynamic card roles -> weighted archetype tag last")
        self.assertGreaterEqual(len(plan["cards"]), 6)
        roles = [item["role"] for item in plan["cards"]]
        self.assertEqual(roles[0], "hook")
        self.assertIn(roles[-1], {"watch", "impact"})
        self.assertEqual(plan["archetype_tag"], "business_transformation")

        pipeline_source = (ROOT / "apps" / "streamlit" / "story_content_pipeline_v5.py").read_text(encoding="utf-8")
        self.assertNotIn("Riot Platforms", pipeline_source)
        self.assertNotIn("Anthropic", pipeline_source)
        self.assertNotIn("Coinspeaker", pipeline_source)
        self.assertNotIn("if archetype ==", pipeline_source)

    def test_local_japanese_mode_prefers_japanese_evidence(self) -> None:
        result = story_content_pipeline_v5.generate_story_package(
            [SAYLOR_STYLE, RIOT_MIXED_LANGUAGE],
            7,
            {"provider": story_content_pipeline_v5.PROVIDER_LOCAL},
            generation_seed="locale-v103",
        )
        self.assertIsNone(result.error)
        package = result.package
        self.assertEqual(package["content_quality"]["hero_selection_reason"], "local_provider_prefers_japanese_evidence")
        self.assertEqual(package["story_context"]["hero_resource_ids"], [RIOT_MIXED_LANGUAGE["id"]])
        cards = package["cards"]["STORY"][:-1]
        self.assertTrue(all(story_source_engine_v5.japanese_ratio(f"{c['headline']} {c['key_message']}") >= 0.35 for c in cards))

    def test_english_only_local_ja_generation_is_blocked(self) -> None:
        result = story_content_pipeline_v5.generate_story_package(
            [SAYLOR_STYLE], 7, {"provider": story_content_pipeline_v5.PROVIDER_LOCAL}, generation_seed="english-block"
        )
        self.assertIsNotNone(result.error)
        self.assertIn("Japanese Story generation blocked", result.error)

    def test_different_resource_type_builds_different_plan_without_code_change(self) -> None:
        result = story_content_pipeline_v5.generate_story_package(
            [POLICY], 7, {"provider": story_content_pipeline_v5.PROVIDER_LOCAL}, generation_seed="policy-v10"
        )
        self.assertIsNone(result.error)
        plan = result.package["story_context"]["story_plan"]
        self.assertEqual(plan["archetype_tag"], "policy_change")
        roles = [item["role"] for item in plan["cards"]]
        self.assertIn("watch", roles)
        self.assertNotIn("old_business", roles)

    def test_pipeline_builds_fact_bound_cards_and_graph_export(self) -> None:
        result = story_content_pipeline_v5.generate_story_package(
            [TRANSFORMATION], 7, {"provider": story_content_pipeline_v5.PROVIDER_LOCAL}, generation_seed="transform-v103"
        )
        self.assertIsNone(result.error)
        package = result.package
        self.assertEqual(package["content_quality"]["pipeline"], "story-content-v10.3")
        self.assertEqual(package["content_quality"]["source_engine"], "story-source-v10.3")
        self.assertEqual(package["content_quality"]["graph_engine"], "story-graph-v10.3")
        cards = package["cards"]["STORY"][:-1]
        self.assertEqual(len(cards), 6)
        self.assertTrue(all((c.get("qa") or {}).get("fact_bound") for c in cards))
        self.assertTrue(all((c.get("qa") or {}).get("claim_evidence_consistent") for c in cards))
        self.assertTrue(all((c.get("qa") or {}).get("evidence_sentence_complete") for c in cards))
        self.assertTrue(all((c.get("qa") or {}).get("scene_evidence_consistent") for c in cards))
        self.assertTrue(all(c.get("evidence_excerpt") for c in cards))
        self.assertGreaterEqual(len({(c.get("visual_direction") or {}).get("scene_type") for c in cards}), 4)
        diag = story_renderer_v5.scene_diagnostics(cards)
        self.assertGreaterEqual(diag["render_signature_count"], 4)

        raw = mode_exporter_v5.build_story_excel(package, [TRANSFORMATION])
        wb = load_workbook(BytesIO(raw))
        self.assertIn("Story_Graph", wb.sheetnames)
        self.assertIn("Story_Plan", wb.sheetnames)
        graph_headers = [cell.value for cell in wb["Story_Graph"][1]]
        self.assertIn("value_details", graph_headers)
        self.assertIn("complete", graph_headers)
        self.assertEqual(len(getattr(wb["Card_Previews"], "_images", [])), 7)

    def test_runtime_routes_to_v10_generic_stack(self) -> None:
        entry = (ROOT / "streamlit_app.py").read_text(encoding="utf-8")
        self.assertIn('RUNTIME_TOKEN = "dual-pipeline-v10.3"', entry)
        self.assertIn('sys.modules["story_engine"] = story_source_engine_v5', entry)
        self.assertIn('sys.modules["story_content_pipeline"] = story_content_pipeline_v5', entry)
        self.assertIn('sys.modules["mode_exporter"] = mode_exporter_v5', entry)
        self.assertIn("story_graph_engine", entry)
        self.assertIn("story_renderer_v5", entry)


if __name__ == "__main__":
    unittest.main()
