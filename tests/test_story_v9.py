from __future__ import annotations

import sys
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "apps" / "streamlit"))

import mode_exporter_v4  # noqa: E402
import story_article_cleaner  # noqa: E402
import story_content_pipeline_v3  # noqa: E402
import story_content_pipeline_v4  # noqa: E402
import story_engine_v3  # noqa: E402
import story_engine_v4  # noqa: E402
import story_renderer_v3  # noqa: E402
import story_renderer_v4  # noqa: E402


RIOT = {
    "id": "riot-1941",
    "source": "Coinspeaker JP",
    "source_type": "rss",
    "region": "Japan/Global",
    "category": "Japanese crypto and finance news",
    "title": "BTCマイナー企業Riot、Anthropicとデータセンター契約か",
    "url": "https://example.com/riot-ai",
    "tags": "BTC, AI, MINING",
    "trader_score": 54,
    "risk_score": 18,
    "excerpt": "Riot PlatformsがBTCマイニングからAIデータセンターへ事業を多角化する。",
    "material": (
        "暗号資産マイニング企業のRiot Platformsは、AI企業Anthropic向けとされる大規模データセンター契約を締結した。 "
        "テキサス州ロックデールで191メガワットのIT容量を提供する。 "
        "契約期間は20年間で、初期契約価値は約91億ドル、延長時は約161億ドルとなる可能性がある。 "
        "Riot Platformsはこれまでビットコインのマイニングを主力事業としてきた。 "
        "既存の電力インフラをAIデータセンターへ転用し、2027年に初期稼働、2028年にフル稼働を計画する。 "
        "Jason Les CEOはデジタルインフラ企業への転換を進めるとした。"
    ),
}

CONTAMINATED_RIOT = {
    **RIOT,
    "id": "riot-dirty",
    "material": (
        "BTCマイナー企業Riot、Anthropicとデータセンター契約か。 Crypto.comの評判 1 BTC = ... "
        "By 黒川 理佐 Editor 倉元 大智 8月 12, 2026 at 08:34 AM Updated 8月 12, 2026 at 08:34 AM 1 min read。 "
        "暗号資産マイニング企業のRiot Platformsは、AI企業Anthropic向けの大規模データセンター契約を締結した。 "
        "テキサス州ロックデールで191メガワットのIT容量を提供する。契約期間は2048年6月までの20年間となっている。 "
        "初期契約価値は約91億ドル、延長時は約161億ドルとなる可能性がある。 "
        "施設の稼働は段階的に進められ、2027年12月までに96メガワット、2028年6月には191メガワットのフル稼働を計画している。 "
        "Riot Platformsはこれまでビットコインのマイニングを主力事業としてきた。既存の電力インフラをAIデータセンターへ転用する。 "
        "初期開発費用として約5億7300万ドルのつなぎ融資も確保している。 "
        "Jason Les CEOはデジタルインフラ企業への転換を進めるとした。 "
        "Disclaimer: Coinspeakerは情報提供を目的としており、投資助言ではありません。 "
        "2021年から仮想通貨投資を始め、2025年よりCoinspeaker参画。 "
        "韓国モバイル大手カカオペイ、ウォン建ステーブルコイン発行準備 ビットフライヤー、BTCハードフォークの第一報を発表 Reproduction in whole or in part in any form or medium without express written permission of Coinspeaker LTD is prohibited."
    ),
}

BITFLYER = {
    "id": "bitflyer-1941",
    "source": "Coinspeaker JP",
    "source_type": "rss",
    "region": "Japan",
    "category": "Japanese crypto and finance news",
    "title": "bitFlyer、ビットコインのハードフォーク新通貨ECXへの対応方針を公表",
    "url": "https://example.com/bitflyer-fork",
    "tags": "BTC, EXCHANGE",
    "trader_score": 48,
    "risk_score": 12,
    "excerpt": "bitFlyerはビットコインから新通貨ECXが分岐する可能性について対応方針を説明した。",
    "material": "bitFlyerはビットコインのハードフォークに伴う新通貨ECXの取扱い方針を公表した。顧客資産の扱いと付与条件を案内する。",
}

HAYES = {
    "id": "hayes-1941",
    "source": "Coinspeaker JP",
    "source_type": "rss",
    "region": "Global",
    "category": "Japanese crypto and finance news",
    "title": "Arthur Hayes、AIバブル崩壊とビットコインの未来を予測",
    "url": "https://example.com/hayes-ai",
    "tags": "BTC, AI, MACRO",
    "trader_score": 50,
    "risk_score": 16,
    "excerpt": "Arthur HayesはAI関連株の調整とビットコイン流動性について見通しを語った。",
    "material": "Arthur HayesはAI投資ブームが調整する可能性と、その後の流動性環境がビットコインへ与える影響について見解を示した。",
}


class StoryV9Test(unittest.TestCase):
    def test_same_publisher_different_events_do_not_cluster(self) -> None:
        rows = story_engine_v4.annotate_resources([RIOT, BITFLYER, HAYES])
        riot = next(r for r in rows if r["id"] == "riot-1941")
        bitflyer = next(r for r in rows if r["id"] == "bitflyer-1941")
        hayes = next(r for r in rows if r["id"] == "hayes-1941")
        self.assertEqual(story_engine_v4.event_similarity(riot, bitflyer), 0.0)
        self.assertEqual(story_engine_v4.event_similarity(riot, hayes), 0.0)
        clusters = story_engine_v4.cluster_story_candidates([RIOT, BITFLYER, HAYES])
        self.assertEqual(sorted(len(c) for c in clusters), [1, 1, 1])

    def test_article_cleaner_removes_coinspeaker_boilerplate(self) -> None:
        cleaned = story_article_cleaner.clean_story_resource(CONTAMINATED_RIOT)
        material = cleaned["material"]
        self.assertNotIn("Crypto.comの評判", material)
        self.assertNotIn("Editor 倉元", material)
        self.assertNotIn("Disclaimer", material)
        self.assertNotIn("Coinspeaker参画", material)
        self.assertNotIn("ビットフライヤー", material)
        self.assertNotIn("Reproduction in whole", material)
        self.assertIn("2027年", material)
        self.assertIn("2028年", material)
        self.assertIn("約91億ドル", material)
        self.assertIn("約161億ドル", material)

    def test_dirty_article_fact_pack_is_clean_and_complete(self) -> None:
        result = story_content_pipeline_v4.generate_story_package(
            [CONTAMINATED_RIOT], 7, {"provider": story_content_pipeline_v4.PROVIDER_LOCAL}, generation_seed="v92-dirty"
        )
        self.assertIsNone(result.error)
        package = result.package
        hero = package["story_context"]["hero_story"]
        pack = package["story_context"]["evidence_facts"]
        blob = str({"hero": hero, "pack": pack, "cards": package["cards"]["STORY"]})
        self.assertNotIn("bitFlyer", blob)
        self.assertNotIn("ビットフライヤー", blob)
        self.assertNotIn("Disclaimer", blob)
        self.assertNotIn("Coinspeaker参画", blob)
        self.assertNotIn("Editor 倉元", blob)
        self.assertIn("2027", pack["milestone_years"])
        self.assertIn("2028", pack["milestone_years"])
        values = set(pack["values"])
        self.assertIn("191メガワット", values)
        self.assertIn("96メガワット", values)
        self.assertIn("約91億ドル", values)
        self.assertIn("約161億ドル", values)
        self.assertIn("約5億7300万ドル", values)
        qa = package["content_quality"]["story_qa"]
        self.assertTrue(qa["article_cleaning_pass"])
        self.assertEqual(qa["boilerplate_fact_count"], 0)
        self.assertEqual(qa["claim_evidence_mismatch_count"], 0)
        self.assertTrue(qa["publishable"])
        watch = next(card for card in package["cards"]["STORY"] if card.get("story_role") == "watch")
        self.assertNotIn("2026", watch["headline"] + " " + watch["key_message"])
        self.assertIn("2027", watch["headline"] + " " + watch["key_message"])
        self.assertIn("2028", watch["headline"] + " " + watch["key_message"])

    def test_hero_cluster_and_evidence_are_riot_only(self) -> None:
        result = story_content_pipeline_v4.generate_story_package(
            [RIOT, BITFLYER, HAYES],
            7,
            {"provider": story_content_pipeline_v4.PROVIDER_LOCAL},
            generation_seed="v9-isolation",
        )
        self.assertIsNone(result.error)
        package = result.package
        hero = package["story_context"]["hero_story"]
        self.assertEqual(hero["archetype"], "business_transformation")
        self.assertEqual(set(hero["resource_ids"]), {"riot-1941"})
        self.assertEqual(set(package["story_context"]["evidence_facts"]["source_ids"]), {"riot-1941"})
        cards = package["cards"]["STORY"][:-1]
        self.assertTrue(all(card["evidence_refs"] == ["riot-1941"] for card in cards))
        hero_output = str({
            "hero": hero,
            "evidence_facts": package["story_context"]["evidence_facts"],
            "cards": package["cards"]["STORY"],
        })
        self.assertNotIn("ECX", hero_output)
        self.assertNotIn("ハードフォーク", hero_output)
        self.assertNotIn("Arthur Hayes", hero_output)

    def test_pixel_level_scene_qa_is_not_metadata_only(self) -> None:
        result = story_content_pipeline_v4.generate_story_package(
            [RIOT], 7, {"provider": story_content_pipeline_v4.PROVIDER_LOCAL}, generation_seed="v9-visual"
        )
        cards = result.package["cards"]["STORY"][:-1]
        diag = story_renderer_v4.scene_diagnostics(cards)
        self.assertGreaterEqual(diag["render_signature_count"], 4)
        self.assertEqual(diag["near_duplicate_scene_pairs"], [])
        qa = result.package["content_quality"]["story_qa"]
        self.assertGreaterEqual(qa["render_signature_count"], 4)
        self.assertTrue(qa["hero_cluster_same_event"])
        self.assertTrue(qa["publishable"])

    def test_excel_contains_v92_qa_and_previews(self) -> None:
        result = story_content_pipeline_v4.generate_story_package(
            [RIOT], 7, {"provider": story_content_pipeline_v4.PROVIDER_LOCAL}, generation_seed="v92-excel"
        )
        raw = mode_exporter_v4.build_story_excel(result.package, [RIOT])
        wb = load_workbook(BytesIO(raw))
        self.assertEqual(len(getattr(wb["Card_Previews"], "_images", [])), 7)
        self.assertIn("Story_QA", wb.sheetnames)
        context_values = [wb["Story_Context"].cell(row=i, column=2).value for i in range(1, wb["Story_Context"].max_row + 1)]
        self.assertTrue(any("story-content-v9.2" in str(v) for v in context_values))
        qa_values = {wb["Story_QA"].cell(row=i, column=1).value: wb["Story_QA"].cell(row=i, column=2).value for i in range(2, wb["Story_QA"].max_row + 1)}
        self.assertIn("render_signature_count", qa_values)
        self.assertIn("cluster_coherence", qa_values)
        self.assertIn("claim_evidence_mismatch_count", qa_values)
        self.assertIn("article_cleaning_pass", qa_values)

    def test_v9_legacy_dependencies_are_real_v3_modules(self) -> None:
        self.assertIs(story_engine_v4.legacy, story_engine_v3)
        self.assertIs(story_content_pipeline_v4.legacy, story_content_pipeline_v3)
        self.assertIs(story_renderer_v4.legacy, story_renderer_v3)
        self.assertIsNot(story_engine_v4.legacy, story_engine_v4)
        self.assertIsNot(story_content_pipeline_v4.legacy, story_content_pipeline_v4)
        self.assertIsNot(story_renderer_v4.legacy, story_renderer_v4)

    def test_runtime_routes_to_v92_without_poisoning_v3_aliases(self) -> None:
        entry = (ROOT / "streamlit_app.py").read_text(encoding="utf-8")
        self.assertIn('RUNTIME_TOKEN = "dual-pipeline-v9.2"', entry)
        self.assertIn('sys.modules["story_engine"] = story_engine_v4', entry)
        self.assertIn('sys.modules["story_content_pipeline"] = story_content_pipeline_v4', entry)
        self.assertIn('sys.modules["mode_exporter"] = mode_exporter_v4', entry)
        self.assertIn('sys.modules["story_engine_v3"] = story_engine_legacy', entry)
        self.assertIn('sys.modules["story_content_pipeline_v3"] = story_content_pipeline_legacy', entry)
        self.assertIn('sys.modules["story_renderer_v3"] = story_renderer_legacy', entry)
        self.assertNotIn('sys.modules["story_engine_v3"] = story_engine_v4', entry)
        self.assertNotIn('sys.modules["story_content_pipeline_v3"] = story_content_pipeline_v4', entry)
        self.assertNotIn('sys.modules["story_renderer_v3"] = story_renderer_v4', entry)


if __name__ == "__main__":
    unittest.main()
