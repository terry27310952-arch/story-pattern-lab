from __future__ import annotations

import re
import sys
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "apps" / "streamlit"))

import card_renderer  # noqa: E402
import content_modes  # noqa: E402
import mode_exporter  # noqa: E402
import mode_resource_pipeline  # noqa: E402
import story_content_pipeline  # noqa: E402
import story_output_guard  # noqa: E402
import story_render_runtime  # noqa: E402
import visual_variation_runtime  # noqa: E402


STORY_RESOURCE = {
    "id": "story-flow-1",
    "source": "Blockworks",
    "source_type": "rss",
    "region": "Global",
    "category": "Institutional crypto and macro",
    "title": "BlackRock Bitcoin ETF inflows hit a new record while BTC price stays range-bound",
    "url": "https://example.com/story-flow",
    "excerpt": "BlackRock and other institutions recorded $2.4 billion in Bitcoin ETF inflows. The buying continued for several sessions, yet Bitcoin price stayed below resistance, creating a visible gap between fund demand and price response.",
    "material": (
        "BlackRock and other institutions recorded $2.4 billion in Bitcoin ETF inflows. "
        "The buying continued for several sessions, yet Bitcoin price stayed range-bound. "
        "Institutional demand, ETF flows, fund holdings and the gap between capital inflow and spot price became the focus. "
    ) * 8,
    "tags": "BTC, ETF",
    "trader_score": 70,
    "risk_score": 18,
}

GENERIC_RESOURCE = {
    "id": "generic-1",
    "source": "Generic Feed",
    "source_type": "rss",
    "region": "Global",
    "category": "Crypto news",
    "title": "Bitcoin market update",
    "url": "https://example.com/generic",
    "excerpt": "Bitcoin traded during the session with ordinary volatility.",
    "material": "Bitcoin traded during the session with ordinary volatility.",
    "tags": "BTC",
    "trader_score": 88,
    "risk_score": 18,
}


class DualModeV6Test(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        visual_variation_runtime.apply_renderer_patch(card_renderer)
        story_render_runtime.apply_renderer_patch(card_renderer)
        story_output_guard.apply_generation_guard(story_content_pipeline)

    def test_source_presets_are_mode_specific(self) -> None:
        public_story = mode_resource_pipeline.available_public_registry(content_modes.MODE_STORY)
        story_rss, story_public = content_modes.default_sources(
            content_modes.MODE_STORY,
            mode_resource_pipeline.RSS_SOURCES,
            public_story,
        )
        trader_rss, trader_public = content_modes.default_sources(
            content_modes.MODE_TRADER,
            mode_resource_pipeline.RSS_SOURCES,
            mode_resource_pipeline.PUBLIC_LIST_SOURCES,
        )
        self.assertIn("Japan FSA Crypto Policy", story_public)
        self.assertIn("SEC Press Releases", story_public)
        self.assertIn("CFTC Press Releases", story_public)
        self.assertNotIn("5ch Crypto Board", story_public)
        self.assertIn("5ch Crypto Board", trader_public)
        self.assertNotEqual(story_rss, trader_rss)

    def test_story_score_can_outrank_higher_trader_score(self) -> None:
        ranked = content_modes.rank_resources(
            content_modes.MODE_STORY,
            [GENERIC_RESOURCE, STORY_RESOURCE],
        )
        self.assertEqual(ranked[0]["id"], STORY_RESOURCE["id"])
        self.assertGreater(ranked[0]["story_score"], ranked[1]["story_score"])

    def test_story_generation_has_no_trader_intermediate(self) -> None:
        result = story_content_pipeline.generate_story_package(
            [STORY_RESOURCE, GENERIC_RESOURCE],
            total_card_count=7,
            config={"provider": story_content_pipeline.PROVIDER_LOCAL},
            output_locale="ja-JP",
            brand={"cta": "フォローして、勢力が入ったポイントを無料でチェック。"},
            generation_seed="dual-mode-test-seed",
        )
        self.assertIsNone(result.error)
        package = result.package
        self.assertEqual(package["mode"], "story")
        self.assertNotIn("trader_stance", package)
        self.assertNotIn("market_structure", package)
        self.assertEqual((package.get("content_quality") or {}).get("pipeline"), "story-content-v6.0")
        self.assertEqual((package.get("content_quality") or {}).get("output_guard"), story_output_guard.STORY_OUTPUT_GUARD_VERSION)

        cards = package["cards"]["스토리"]
        self.assertEqual(len(cards), 7)
        self.assertEqual(cards[-1]["card_type"], "brand_outro")
        self.assertEqual(cards[-1]["headline"], "勢力ハンター キヨサキ")
        self.assertEqual(cards[-1]["eyebrow"], "キヨサキ")
        lock = cards[-1]["visual_direction"]["character_style_lock"]
        self.assertIn("featureless", lock["face"])
        self.assertIn("black leather gloves", lock["wardrobe"])

        roles = []
        for card in cards[:-1]:
            roles.append(card.get("story_role"))
            direction = card.get("visual_direction") or {}
            self.assertFalse(direction.get("character_required"))
            self.assertEqual(direction.get("character_visibility"), 0.0)
            visible = " ".join(str(card.get(key) or "") for key in ["eyebrow", "headline", "subheadline", "key_message"])
            self.assertNotRegex(visible, r"[가-힣]")
            self.assertNotIn("THE OBSERVER", visible.upper())
            prompts = direction.get("image_prompts") or {}
            self.assertIn("No K monogram", prompts.get("4:5", ""))
        self.assertGreater(len(set(roles)), 3)

    def test_story_layout_sequence_changes_across_generation_seeds(self) -> None:
        sequences = set()
        for seed in ["A", "B", "C", "D", "E"]:
            result = story_content_pipeline.generate_story_package(
                [STORY_RESOURCE, GENERIC_RESOURCE],
                total_card_count=7,
                config={"provider": story_content_pipeline.PROVIDER_LOCAL},
                generation_seed=seed,
            )
            cards = result.package["cards"]["스토리"][:-1]
            sequences.add(tuple((card.get("visual_direction") or {}).get("layout_variant") for card in cards))
        self.assertGreater(len(sequences), 1)

    def test_story_excel_embeds_same_renderer_previews(self) -> None:
        result = story_content_pipeline.generate_story_package(
            [STORY_RESOURCE, GENERIC_RESOURCE],
            total_card_count=6,
            config={"provider": story_content_pipeline.PROVIDER_LOCAL},
            generation_seed="excel-story",
        )
        raw = mode_exporter.build_story_excel(result.package, [STORY_RESOURCE, GENERIC_RESOURCE])
        wb = load_workbook(BytesIO(raw))
        for sheet in ["Card_Previews", "Story_Context", "Story_Candidates", "Story_Cards", "Sources", "Note"]:
            self.assertIn(sheet, wb.sheetnames)
        self.assertGreaterEqual(len(getattr(wb["Card_Previews"], "_images", [])), 6)

    def test_production_entrypoint_does_not_apply_story_to_trader_reasoning(self) -> None:
        entry = (ROOT / "streamlit_app.py").read_text(encoding="utf-8")
        self.assertIn('APP_FILE = APP_DIR / "app_v2.py"', entry)
        self.assertNotIn("story_pipeline_runtime.apply_reasoning_patch", entry)
        self.assertNotIn("story_deck_runtime.apply_reasoning_patch", entry)
        self.assertNotIn("visual_variation_runtime.apply_reasoning_patch", entry)
        self.assertIn("apply_brand_patch(reasoning_engine)", entry)
        self.assertIn("story_output_guard.apply_generation_guard", entry)

    def test_removed_reference_asset_stays_removed(self) -> None:
        self.assertFalse((ROOT / "assets" / "brand" / "observer_reference.png").exists())


if __name__ == "__main__":
    unittest.main()
