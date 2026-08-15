from __future__ import annotations

import sys
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "apps" / "streamlit"))

import mode_exporter_v3  # noqa: E402
import story_content_pipeline_v3  # noqa: E402
import story_engine_v3  # noqa: E402
import story_renderer_v3  # noqa: E402


RIOT = {
    "id": "riot-1",
    "source": "Coinspeaker JP",
    "source_type": "rss",
    "region": "Japan/Global",
    "category": "Japanese crypto and finance news",
    "title": "BTCマイナー企業Riot、アンソロピックとデータセンター契約か",
    "url": "https://example.com/riot",
    "tags": "BTC, AI, MINING",
    "trader_score": 54,
    "risk_score": 18,
    "excerpt": "Riot PlatformsがBTCマイニングからAIデータセンターへ事業を多角化する。",
    "material": (
        "暗号資産マイニング企業のRiot Platformsは、最先端AI企業と大規模なデータセンターのリース契約を締結した。 "
        "報道によると契約先はAnthropicだという。テキサス州ロックデールで191メガワットのIT容量を提供する。 "
        "契約期間は20年間で、初期契約による収益は約91億ドルに達する見込みだ。 "
        "延長オプションをすべて行使した場合、総契約価値は約161億ドルとなる可能性がある。 "
        "Riot Platformsはこれまでビットコインのマイニングを主力事業としてきた。 "
        "今後は既存の電力インフラを活かし、AIデータセンター開発への多角化を進める。 "
        "2027年12月までに初期96メガワット、2028年6月には191メガワットのフル稼働を計画する。 "
        "同社は初期開発費用として約5億7300万ドルのつなぎ融資も確保している。 "
        "ジェイソン・レスCEOは仮想通貨とAIの両分野でデジタルインフラのリーダーを目指すとした。"
    ),
}

HISTORY = {
    "id": "history-1",
    "source": "BeInCrypto",
    "source_type": "rss",
    "region": "Global",
    "category": "Markets",
    "title": "Wall Street Only Looked Like This in 1929 and 2000 as Shiller CAPE nears extremes",
    "url": "https://example.com/history",
    "tags": "MACRO, VALUATION",
    "trader_score": 50,
    "risk_score": 18,
    "excerpt": "Shiller CAPE is near levels associated with 1929 and 2000.",
    "material": (
        "The Shiller CAPE ratio is near 40 to 42. The dot-com peak in 2000 reached roughly 44. "
        "Analysts compare the current valuation with 1929 and 2000, while noting that liquidity and market structure differ today."
    ),
}


class StoryV8Test(unittest.TestCase):
    def test_riot_is_business_transformation(self) -> None:
        row = story_engine_v3.annotate_resource(RIOT)
        self.assertEqual(row["story_archetype_hint"], "business_transformation")
        self.assertIn("Riot Platforms", row["story_entities"])
        self.assertIn("Anthropic", row["story_entities"])
        self.assertNotIn("Crypto.", row["story_entities"])
        self.assertNotIn("Editor", row["story_entities"])

    def test_hero_fact_pack_does_not_cross_candidates(self) -> None:
        result = story_content_pipeline_v3.generate_story_package(
            [RIOT, HISTORY],
            total_card_count=7,
            config={"provider": story_content_pipeline_v3.PROVIDER_LOCAL},
            generation_seed="riot-isolation",
        )
        self.assertIsNone(result.error)
        package = result.package
        hero = package["story_context"]["hero_story"]
        self.assertEqual(hero["archetype"], "business_transformation")
        self.assertEqual(set(package["story_context"]["evidence_facts"]["source_ids"]), {"riot-1"})
        evidence_blob = str(package["story_context"]["evidence_facts"])
        self.assertNotIn("1929", evidence_blob)
        self.assertNotIn("2000", evidence_blob)
        self.assertNotIn("CAPE", evidence_blob.upper())
        cards = package["cards"]["STORY"][:-1]
        self.assertTrue(all(set(card.get("evidence_refs") or []).issubset({"riot-1"}) for card in cards))
        visible = " ".join((card.get("headline") or "") + " " + (card.get("key_message") or "") for card in cards)
        self.assertIn("Riot", visible)
        self.assertTrue("191" in visible or "91億" in visible)

    def test_history_is_isolated_when_used_alone(self) -> None:
        result = story_content_pipeline_v3.generate_story_package(
            [HISTORY], 7, {"provider": story_content_pipeline_v3.PROVIDER_LOCAL}, generation_seed="history"
        )
        self.assertEqual(result.package["story_context"]["hero_story"]["archetype"], "historical_parallel")
        blob = str(result.package)
        self.assertIn("1929", blob)
        self.assertIn("2000", blob)
        self.assertIn("CAPE", blob.upper())
        self.assertNotIn("Anthropic", blob)
        self.assertNotIn("Riot Platforms", blob)

    def test_story_has_distinct_scenes_and_layouts(self) -> None:
        result = story_content_pipeline_v3.generate_story_package(
            [RIOT], 7, {"provider": story_content_pipeline_v3.PROVIDER_LOCAL}, generation_seed="visual"
        )
        cards = result.package["cards"]["STORY"][:-1]
        scenes = {(c.get("visual_direction") or {}).get("scene_type") for c in cards}
        layouts = {(c.get("visual_direction") or {}).get("layout_variant") for c in cards}
        self.assertGreaterEqual(len(scenes), 4)
        self.assertGreaterEqual(len(layouts), 4)
        pngs = [story_renderer_v3.render_story_card_png(card, width=360, height=450) for card in cards]
        self.assertTrue(all(png.startswith(b"\x89PNG") for png in pngs))
        self.assertGreaterEqual(len({hash(png) for png in pngs}), 4)

    def test_story_qa_publishable_and_excel_embeds_previews(self) -> None:
        result = story_content_pipeline_v3.generate_story_package(
            [RIOT], 7, {"provider": story_content_pipeline_v3.PROVIDER_LOCAL}, generation_seed="excel"
        )
        qa = result.package["content_quality"]["story_qa"]
        self.assertTrue(qa["hero_evidence_isolated"])
        self.assertEqual(qa["generic_card_count"], 0)
        self.assertGreaterEqual(qa["unique_scene_types"], 4)
        self.assertGreaterEqual(qa["unique_layouts"], 4)
        self.assertTrue(qa["publishable"])
        raw = mode_exporter_v3.build_story_excel(result.package, [RIOT])
        wb = load_workbook(BytesIO(raw))
        self.assertIn("Story_QA", wb.sheetnames)
        self.assertIn("Card_Previews", wb.sheetnames)
        self.assertEqual(len(getattr(wb["Card_Previews"], "_images", [])), 7)

    def test_runtime_routes_story_layers_to_v9(self) -> None:
        entry = (ROOT / "streamlit_app.py").read_text(encoding="utf-8")
        self.assertIn('RUNTIME_TOKEN = "dual-pipeline-v9.0"', entry)
        self.assertIn('sys.modules["story_engine"] = story_engine_v4', entry)
        self.assertIn('sys.modules["story_content_pipeline"] = story_content_pipeline_v4', entry)
        self.assertIn('sys.modules["mode_exporter"] = mode_exporter_v4', entry)


if __name__ == "__main__":
    unittest.main()
