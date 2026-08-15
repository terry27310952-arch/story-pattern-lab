from __future__ import annotations

import sys
import unittest
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "apps" / "streamlit"))

import card_renderer  # noqa: E402
import story_engine  # noqa: E402
import story_export_runtime  # noqa: E402
import story_pipeline_runtime  # noqa: E402
import story_render_runtime  # noqa: E402
import visual_variation_runtime  # noqa: E402


MONEY_FLOW = {
    "id": "flow-1",
    "source": "Blockworks",
    "source_type": "rss",
    "title": "BlackRock Bitcoin ETF inflows hit record as institutions keep buying",
    "url": "https://example.com/flow",
    "excerpt": "Institutional funds recorded $2.4 billion of Bitcoin ETF inflows, yet spot price remained below resistance. The fund has expanded its holdings for several sessions.",
    "material": "Institutional funds recorded $2.4 billion of Bitcoin ETF inflows, yet spot price remained below resistance. BlackRock and other funds continued buying while Bitcoin price stayed range-bound. The change is drawing attention to whether persistent demand eventually reaches price.",
    "tags": "BTC, ETF",
    "trader_score": 72,
    "risk_score": 18,
}

GENERIC = {
    "id": "generic-1",
    "source": "Feed",
    "source_type": "rss",
    "title": "Bitcoin market update",
    "url": "https://example.com/generic",
    "excerpt": "Bitcoin traded during the session.",
    "material": "Bitcoin traded during the session with normal volatility.",
    "tags": "BTC",
    "trader_score": 78,
    "risk_score": 18,
}

DUPLICATE_FLOW = {
    **MONEY_FLOW,
    "id": "flow-2",
    "source": "CoinDesk Global",
    "url": "https://example.com/flow-2",
    "title": "BlackRock Bitcoin ETF sees record inflows while BTC price stays range-bound",
}

BASE_CARDS = [
    {
        "slide": 1,
        "card_id": "c1",
        "card_type": "market_conclusion",
        "headline": "センチメントは弱い。",
        "subheadline": "価格はまだ崩れていない。",
        "key_message": "価格を見る。",
        "metrics": [{"id": "btc_price", "label": "BTC", "value": "$63,000", "raw_value": 63000, "locked": True}],
        "source": {},
        "qa": {"renderable": True},
        "visual_direction": {"image_prompts": {"4:5": "base", "9:16": "base"}},
    },
    {
        "slide": 2,
        "card_id": "c2",
        "card_type": "news_context",
        "headline": "ニュースを見る。",
        "subheadline": "材料と価格を分ける。",
        "key_message": "材料と価格を分ける。",
        "metrics": [{"id": "btc_price", "label": "BTC", "value": "$63,000", "raw_value": 63000, "locked": True}],
        "source": {},
        "qa": {"renderable": True},
        "visual_direction": {"image_prompts": {"4:5": "base", "9:16": "base"}},
    },
    {
        "slide": 3,
        "card_id": "c3",
        "card_type": "trade_plan",
        "headline": "次を見る。",
        "subheadline": "条件を確認。",
        "key_message": "条件を確認。",
        "metrics": [{"id": "btc_price", "label": "BTC", "value": "$63,000", "raw_value": 63000, "locked": True}],
        "source": {},
        "qa": {"renderable": True},
        "visual_direction": {"image_prompts": {"4:5": "base", "9:16": "base"}},
    },
    {
        "slide": 4,
        "card_id": "outro",
        "card_type": "brand_outro",
        "headline": "勢力ハンター キヨサキ",
        "subheadline": "The Observer",
        "key_message": "フォローして、勢力が入ったポイントを無料でチェック。",
        "metrics": [],
        "source": {},
        "qa": {"renderable": True},
        "visual_direction": {"image_prompts": {"4:5": "base", "9:16": "base"}},
    },
]


class StoryEngineV5Test(unittest.TestCase):
    def test_story_score_can_outrank_higher_trader_score(self) -> None:
        ranked = story_engine.annotate_resources([GENERIC, MONEY_FLOW])
        self.assertEqual(ranked[0]["id"], "flow-1")
        self.assertGreater(ranked[0]["story_score"], ranked[1]["story_score"])
        self.assertEqual(ranked[0]["story_archetype_hint"], "money_flow")

    def test_same_event_is_clustered(self) -> None:
        clusters = story_engine.cluster_story_candidates([MONEY_FLOW, DUPLICATE_FLOW, GENERIC])
        self.assertTrue(any(len(cluster) >= 2 for cluster in clusters))

    def test_archetypes_have_different_story_arcs(self) -> None:
        money = story_engine.story_arc("money_flow", 6)
        policy = story_engine.story_arc("policy_change", 6)
        self.assertNotEqual(money, policy)
        self.assertEqual(money[0], "hook")
        self.assertEqual(money[-1], "watch")

    def test_storyify_removes_character_from_content_but_locks_outro(self) -> None:
        brief = {"story_context": story_engine.story_context([MONEY_FLOW, GENERIC])}
        package = {
            "cards": {"6장": BASE_CARDS},
            "content_quality": {"visual_blueprint": {"id": "seed-1", "family": "documentary"}},
            "note_markdown": "# Note",
        }
        result = story_pipeline_runtime.storyify_package(package, brief, [MONEY_FLOW, GENERIC])
        cards = result["cards"]["6장"]
        for card in cards[:-1]:
            self.assertEqual(card["visual_direction"]["character_visibility"], 0.0)
            self.assertFalse(card["visual_direction"]["character_required"])
            self.assertNotIn("THE OBSERVER", (card.get("headline") or "").upper())
        self.assertEqual(cards[-1]["eyebrow"], "キヨサキ")
        self.assertTrue(cards[-1]["visual_direction"]["character_required"])
        self.assertIn("black leather gloves", cards[-1]["visual_direction"]["character_style_lock"]["wardrobe"])

    def test_story_renderer_changes_visual_by_archetype(self) -> None:
        visual_variation_runtime.apply_renderer_patch(card_renderer)
        story_render_runtime.apply_renderer_patch(card_renderer)
        a = dict(BASE_CARDS[0])
        b = dict(BASE_CARDS[0])
        a["story_id"] = "a"
        a["story_role"] = "hook"
        a["story_archetype"] = "money_flow"
        a["visual_direction"] = {"format_variant": "poster_center"}
        b["story_id"] = "b"
        b["story_role"] = "hook"
        b["story_archetype"] = "policy_change"
        b["visual_direction"] = {"format_variant": "poster_center"}
        png_a = card_renderer.render_card_png(a, width=432, height=540)
        png_b = card_renderer.render_card_png(b, width=432, height=540)
        self.assertTrue(png_a.startswith(b"\x89PNG"))
        self.assertTrue(png_b.startswith(b"\x89PNG"))
        self.assertNotEqual(png_a, png_b)

    def test_excel_patch_adds_story_sheets(self) -> None:
        def base_excel(brief, content_package, resources, market_snapshot):
            wb = Workbook()
            wb.active.title = "Briefing"
            out = BytesIO()
            wb.save(out)
            return out.getvalue()

        fake = SimpleNamespace(build_excel_bytes=base_excel)
        story_export_runtime.apply_excel_patch(fake)
        brief = {"story_context": story_engine.story_context([MONEY_FLOW, GENERIC])}
        package = story_pipeline_runtime.storyify_package({"cards": {"6장": BASE_CARDS}, "content_quality": {}}, brief, [MONEY_FLOW, GENERIC])
        raw = fake.build_excel_bytes(brief, package, [MONEY_FLOW, GENERIC], {})
        wb = load_workbook(BytesIO(raw))
        self.assertIn("Story_Context", wb.sheetnames)
        self.assertIn("Story_Candidates", wb.sheetnames)
        self.assertIn("Story_Cards", wb.sheetnames)


if __name__ == "__main__":
    unittest.main()
