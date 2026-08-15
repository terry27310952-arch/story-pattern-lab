from __future__ import annotations

import sys
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "apps" / "streamlit"))

from card_renderer import render_card_png  # noqa: E402
from excel_preview_sheet import add_card_preview_sheet  # noqa: E402
from preview_runtime import _dedupe_and_editorialize_cards  # noqa: E402

SAMPLE_CARD = {
    "slide": 1,
    "card_type": "key_levels",
    "headline": "まず見るのはこの2点",
    "key_message": "下は$62,497。上は$65,818を回復できるか。",
    "footer": "キヨサキ · Market Data · Canonical snapshot",
    "qa": {"renderable": True},
    "metrics": [
        {"id": "btc_price", "label": "BTC", "value": "$63,042", "raw_value": 63042},
        {"id": "btc_primary_support", "label": "PRIMARY SUPPORT", "value": "$62,497", "raw_value": 62497},
        {"id": "btc_primary_resistance", "label": "PRIMARY RESISTANCE", "value": "$65,818", "raw_value": 65818},
    ],
}


class PreviewRuntimeTest(unittest.TestCase):
    def test_renderer_returns_real_png_without_reference_sheet_dependency(self) -> None:
        png = render_card_png(SAMPLE_CARD, width=540, height=675)
        self.assertTrue(png.startswith(b"\x89PNG"))
        self.assertGreater(len(png), 10000)

    def test_excel_preview_sheet_contains_embedded_images(self) -> None:
        workbook = Workbook()
        package = {"cards": {"자율제안": [SAMPLE_CARD]}}
        embedded = add_card_preview_sheet(workbook, package)
        self.assertEqual(embedded, 1)
        self.assertIn("Card_Previews", workbook.sheetnames)
        self.assertEqual(len(workbook["Card_Previews"]._images), 1)

        payload = BytesIO()
        workbook.save(payload)
        loaded = load_workbook(BytesIO(payload.getvalue()))
        self.assertIn("Card_Previews", loaded.sheetnames)
        self.assertEqual(len(loaded["Card_Previews"]._images), 1)

    def test_excel_runtime_wrapper_guarantees_preview_sheet(self) -> None:
        import excel_exporter
        import excel_export_runtime

        original_builder = excel_exporter.build_excel_bytes
        saved_base = getattr(excel_exporter, "_kiyosaki_excel_export_base_builder", None)
        saved_version = getattr(excel_exporter, "_kiyosaki_excel_export_runtime_version", None)

        def minimal_builder(brief, content_package, resources, market_snapshot):
            workbook = Workbook()
            workbook.active.title = "Briefing"
            output = BytesIO()
            workbook.save(output)
            return output.getvalue()

        try:
            excel_exporter.build_excel_bytes = minimal_builder
            excel_exporter._kiyosaki_excel_export_base_builder = minimal_builder
            if hasattr(excel_exporter, "_kiyosaki_excel_export_runtime_version"):
                delattr(excel_exporter, "_kiyosaki_excel_export_runtime_version")

            excel_export_runtime.apply_excel_export_patch()
            payload = excel_exporter.build_excel_bytes(
                {}, {"cards": {"자율제안": [SAMPLE_CARD]}}, [], {}
            )
            loaded = load_workbook(BytesIO(payload))
            self.assertIn("Card_Previews", loaded.sheetnames)
            self.assertEqual(len(loaded["Card_Previews"]._images), 1)
            self.assertIn("excel-preview-v4.2", str(loaded["Briefing"]["B3"].value))
        finally:
            excel_exporter.build_excel_bytes = original_builder
            if saved_base is None:
                if hasattr(excel_exporter, "_kiyosaki_excel_export_base_builder"):
                    delattr(excel_exporter, "_kiyosaki_excel_export_base_builder")
            else:
                excel_exporter._kiyosaki_excel_export_base_builder = saved_base
            if saved_version is None:
                if hasattr(excel_exporter, "_kiyosaki_excel_export_runtime_version"):
                    delattr(excel_exporter, "_kiyosaki_excel_export_runtime_version")
            else:
                excel_exporter._kiyosaki_excel_export_runtime_version = saved_version

    def test_duplicate_trade_plan_is_collapsed_and_localized(self) -> None:
        cards = [
            {"slide": 1, "card_type": "market_conclusion", "headline": "A"},
            {"slide": 2, "card_type": "trade_plan", "headline": "Risk Control"},
            {
                "slide": 3,
                "card_type": "trade_plan",
                "headline": "Trade Plan",
                "metrics": [
                    {"id": "btc_primary_support", "value": "$62,497"},
                    {"id": "btc_primary_resistance", "value": "$65,818"},
                ],
            },
            {"slide": 4, "card_type": "brand_outro", "headline": "勢力ハンター キヨサキ"},
        ]
        cleaned = _dedupe_and_editorialize_cards(cards)
        trade_cards = [card for card in cleaned if card.get("card_type") == "trade_plan"]
        self.assertEqual(len(trade_cards), 1)
        self.assertEqual(trade_cards[0]["trade_plan"]["entry"]["condition"], "$65,818を終値で回復")
        self.assertEqual(trade_cards[0]["trade_plan"]["wait"]["condition"], "レンジ内は待機")
        self.assertEqual(trade_cards[0]["trade_plan"]["invalid"]["condition"], "$62,497を終値で割る")
        self.assertEqual([card["slide"] for card in cleaned], list(range(1, len(cleaned) + 1)))

    def test_news_headline_uses_editorial_japanese_fallback(self) -> None:
        cards = [
            {
                "slide": 1,
                "card_type": "news_context",
                "headline": "BTC材料を価格で確認",
                "source": {
                    "publisher": "U.Today",
                    "short_title": "Binance's CZ Hints Bitcoin's Supply Is Even Lower Than Expected",
                    "news_reaction": {"available": False},
                },
            }
        ]
        cleaned = _dedupe_and_editorialize_cards(cards)
        self.assertEqual(cleaned[0]["headline"], "BTC供給の希少性が再び焦点に")
        self.assertIn("反応データがない", cleaned[0]["key_message"])


if __name__ == "__main__":
    unittest.main()
