from __future__ import annotations

import re
import sys
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "apps" / "streamlit"))

import mode_exporter_v4  # noqa: E402
import story_content_pipeline_v4  # noqa: E402
import story_engine_v4  # noqa: E402
import story_renderer_v4  # noqa: E402


HISTORY_RESOURCE = {
    "id": "history-valuation-1",
    "source": "Investor Editorial",
    "source_type": "rss",
    "region": "USA",
    "category": "Equities",
    "title": "Wall Street Only Looked Like This in 1929 and 2000 as Shiller CAPE nears historic extremes",
    "url": "https://example.com/history",
    "excerpt": "The Shiller CAPE is around 41.7 today, close to the dot-com peak near 44.2 in 2000. The article compares the current valuation backdrop with 1929 and 2000.",
    "material": (
        "The Shiller CAPE is around 41.7 today, close to the dot-com peak near 44.2 in 2000. "
        "The article compares the current valuation backdrop with 1929 and 2000 because those were rare periods when cyclically adjusted valuations reached historic extremes. "
        "The comparison concerns valuation rather than an identical chart pattern. "
        "Long-term returns historically became more difficult when starting valuations were unusually high. "
    ) * 4,
    "tags": "WALL STREET, VALUATION, HISTORY",
    "trader_score": 62,
    "risk_score": 14,
}

GENERIC_RESOURCE = {
    "id": "generic-1",
    "source": "Generic Feed",
    "source_type": "rss",
    "region": "Global",
    "category": "Crypto",
    "title": "Bitcoin market update",
    "url": "https://example.com/generic",
    "excerpt": "Bitcoin traded during the session with ordinary volatility.",
    "material": "Bitcoin traded during the session with ordinary volatility.",
    "tags": "BTC",
    "trader_score": 90,
    "risk_score": 15,
}


class ProductionStoryV9CompatibilityTest(unittest.TestCase):
    def test_exact_term_matching_prevents_sec_and_ai_false_positives(self) -> None:
        row = story_engine_v4.annotate_resource(HISTORY_RESOURCE)
        self.assertEqual(row["story_archetype_hint"], "historical_parallel")
        self.assertNotIn("SEC", row.get("story_entities") or [])
        motifs = " ".join(row.get("story_visual_motifs") or []).lower()
        self.assertNotIn("data center", motifs)
        self.assertRegex(motifs, r"archiv")

    def test_story_pipeline_is_source_specific_without_external_model(self) -> None:
        result = story_content_pipeline_v4.generate_story_package(
            [HISTORY_RESOURCE, GENERIC_RESOURCE],
            total_card_count=7,
            config={"provider": story_content_pipeline_v4.PROVIDER_LOCAL},
            generation_seed="history-v9",
        )
        self.assertIsNone(result.error)
        self.assertEqual(result.package["mode"], "story")
        self.assertEqual((result.package.get("content_quality") or {}).get("pipeline"), "story-content-v9.2")
        cards = result.package["cards"]["STORY"]
        self.assertEqual(len(cards), 7)
        visible = " ".join(str(card.get("headline") or "") + " " + str(card.get("key_message") or "") for card in cards[:-1])
        self.assertIn("1929", visible)
        self.assertIn("2000", visible)
        self.assertRegex(visible, r"CAPE|シラー")
        self.assertNotRegex(visible, r"[가-힣]")
        self.assertNotIn("ENTRY", visible)
        self.assertNotIn("Funding", visible)
        self.assertTrue(all(card.get("evidence_excerpt") for card in cards[:-1]))

    def test_story_renderer_uses_actual_scene_diagnostics(self) -> None:
        result = story_content_pipeline_v4.generate_story_package(
            [HISTORY_RESOURCE],
            total_card_count=7,
            config={"provider": story_content_pipeline_v4.PROVIDER_LOCAL},
            generation_seed="render-v9",
        )
        cards = result.package["cards"]["STORY"][:-1]
        layouts = [(card.get("visual_direction") or {}).get("layout_variant") for card in cards]
        scenes = [(card.get("visual_direction") or {}).get("scene_type") for card in cards]
        self.assertGreater(len(set(layouts)), 3)
        self.assertGreater(len(set(scenes)), 3)
        diag = story_renderer_v4.scene_diagnostics(cards)
        self.assertGreaterEqual(diag["render_signature_count"], 4)
        pngs = [story_renderer_v4.render_story_card_png(card, width=324, height=405) for card in cards]
        self.assertTrue(all(png.startswith(b"\x89PNG") for png in pngs))

    def test_story_excel_uses_v9_renderer_and_contains_no_korean_output_labels(self) -> None:
        result = story_content_pipeline_v4.generate_story_package(
            [HISTORY_RESOURCE],
            total_card_count=7,
            config={"provider": story_content_pipeline_v4.PROVIDER_LOCAL},
            generation_seed="excel-v9",
        )
        raw = mode_exporter_v4.build_story_excel(result.package, [HISTORY_RESOURCE])
        wb = load_workbook(BytesIO(raw))
        self.assertGreaterEqual(len(getattr(wb["Card_Previews"], "_images", [])), 7)
        self.assertEqual(wb["Story_Cards"]["A2"].value, "STORY")
        self.assertIn("Story_QA", wb.sheetnames)
        visible_cells = []
        for ws_name in ["Card_Previews", "Story_Context", "Story_Cards", "Story_QA"]:
            for row in wb[ws_name].iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        visible_cells.append(cell.value)
        self.assertFalse([value for value in visible_cells if re.search(r"[가-힣]", value)])

    def test_production_entrypoint_routes_to_v9_modules(self) -> None:
        entry = (ROOT / "streamlit_app.py").read_text(encoding="utf-8")
        self.assertIn("story_content_pipeline_v4", entry)
        self.assertIn("mode_exporter_v4", entry)
        self.assertIn("story_renderer_v4", entry)
        self.assertIn('sys.modules["story_engine"] = story_engine_v4', entry)
        self.assertIn('sys.modules["story_content_pipeline"] = story_content_pipeline_v4', entry)
        self.assertNotIn("story_pipeline_runtime.apply_reasoning_patch", entry)


if __name__ == "__main__":
    unittest.main()
