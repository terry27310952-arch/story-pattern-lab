from __future__ import annotations

from io import BytesIO
from math import ceil

from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

from card_renderer import render_card_png


PREVIEW_SHEET_NAME = "Card_Previews"


def _production_cards(cards: list[dict]) -> list[dict]:
    return [card for card in (cards or []) if (card.get("qa") or {}).get("renderable", True)]


def add_card_preview_sheet(workbook, content_package: dict) -> int:
    """Embed the same rendered card PNGs shown in Streamlit into the workbook.

    Returns the number of embedded preview images.
    """
    if PREVIEW_SHEET_NAME in workbook.sheetnames:
        del workbook[PREVIEW_SHEET_NAME]

    ws = workbook.create_sheet(PREVIEW_SHEET_NAME, 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    for column_letter in "ABCDEFGHIJKLMN":
        ws.column_dimensions[column_letter].width = 11

    ws.merge_cells("A1:N1")
    ws["A1"] = "キヨサキ · Card Preview Gallery"
    ws["A1"].fill = PatternFill("solid", fgColor="0A0A0A")
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:N2")
    ws["A2"] = "Streamlit 미리보기와 동일한 card_renderer PNG가 Excel 파일 내부에 직접 삽입됩니다."
    ws["A2"].font = Font(color="666666", italic=True, size=9)
    ws["A2"].alignment = Alignment(vertical="center")

    row_cursor = 4
    embedded = 0

    for set_label, cards in (content_package.get("cards") or {}).items():
        visible_cards = _production_cards(cards)
        if not visible_cards:
            continue

        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=14)
        heading = ws.cell(row=row_cursor, column=1)
        heading.value = f"{set_label} · {len(visible_cards)} cards"
        heading.font = Font(color="E68A19", bold=True, size=11)
        row_cursor += 2

        pair_rows = ceil(len(visible_cards) / 2)
        for index, card in enumerate(visible_cards):
            grid_col = 1 if index % 2 == 0 else 8
            grid_row = row_cursor + (index // 2) * 25

            ws.merge_cells(
                start_row=grid_row,
                start_column=grid_col,
                end_row=grid_row,
                end_column=grid_col + 5,
            )
            label = ws.cell(row=grid_row, column=grid_col)
            label.value = f"{card.get('slide', index + 1)}. {card.get('headline', '')}"
            label.font = Font(bold=True, size=10)
            label.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[grid_row].height = 28

            png = render_card_png(card, width=540, height=675)
            image = XLImage(BytesIO(png))
            image.width = 270
            image.height = 338
            ws.add_image(image, ws.cell(row=grid_row + 1, column=grid_col).coordinate)
            embedded += 1

            for image_row in range(grid_row + 1, grid_row + 24):
                ws.row_dimensions[image_row].height = 16

        row_cursor += pair_rows * 25 + 2

    if embedded == 0:
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor + 2, end_column=14)
        ws.cell(row=row_cursor, column=1).value = "렌더 가능한 카드가 없어 미리보기 이미지를 만들지 못했습니다."
        ws.cell(row=row_cursor, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_cursor, column=1).font = Font(color="999999", italic=True)

    return embedded
