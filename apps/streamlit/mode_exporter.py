from __future__ import annotations

import json
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

import card_renderer
import excel_exporter


MODE_EXPORTER_VERSION = "mode-exporter-v6.0"
HEADER_FILL = PatternFill("solid", fgColor="171A1C")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ACCENT_FILL = PatternFill("solid", fgColor="F69F19")


def _stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _style_table(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in ws.columns:
        letter = column[0].column_letter
        width = min(48, max(12, max((len(_stringify(cell.value)) for cell in column[:80]), default=10) + 2))
        ws.column_dimensions[letter].width = width


def _append_rows(ws, headers: list[str], rows: list[dict]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([_stringify(row.get(key, "")) for key in headers])
    _style_table(ws)


def _renderable_cards(package: dict) -> list[tuple[str, dict]]:
    output: list[tuple[str, dict]] = []
    for set_name, cards in (package.get("cards") or {}).items():
        for card in cards or []:
            if (card.get("qa") or {}).get("renderable", True):
                output.append((str(set_name), card))
    return output


def add_preview_sheet(wb: Workbook, package: dict) -> int:
    if "Card_Previews" in wb.sheetnames:
        del wb["Card_Previews"]
    ws = wb.create_sheet("Card_Previews", 0)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "キヨサキ · Card Previews"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = "이 시트의 이미지는 앱 미리보기와 동일한 card_renderer.render_card_png 결과입니다."
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 42

    streams: list[BytesIO] = []
    count = 0
    row_cursor = 4
    for index, (set_name, card) in enumerate(_renderable_cards(package)):
        col = "A" if index % 2 == 0 else "B"
        if index and index % 2 == 0:
            row_cursor += 22
        anchor = f"{col}{row_cursor + 2}"
        ws[f"{col}{row_cursor}"] = f"{set_name} · {card.get('slide')} · {card.get('headline', '')}"
        ws[f"{col}{row_cursor}"].font = Font(bold=True)
        ws[f"{col}{row_cursor + 1}"] = f"{card.get('story_role') or card.get('card_type') or ''}"
        stream = BytesIO(card_renderer.render_card_png(card, width=540, height=675))
        streams.append(stream)
        image = XLImage(stream)
        image.width = 270
        image.height = 337
        ws.add_image(image, anchor)
        count += 1
    # Keep streams alive through workbook serialization. openpyxl reads them during save.
    wb._kiyosaki_preview_streams = streams  # type: ignore[attr-defined]
    return count


def _save_verified(wb: Workbook, expected_previews: int) -> bytes:
    output = BytesIO()
    wb.save(output)
    raw = output.getvalue()
    verify = load_workbook(BytesIO(raw))
    if "Card_Previews" not in verify.sheetnames:
        raise RuntimeError("Card_Previews sheet missing after Excel serialization")
    actual = len(getattr(verify["Card_Previews"], "_images", []))
    if actual < expected_previews:
        raise RuntimeError(f"Excel preview verification failed: expected {expected_previews}, got {actual}")
    return raw


def build_trader_excel(brief: dict, package: dict, resources: list[dict], market_snapshot: dict) -> bytes:
    base = excel_exporter.build_excel_bytes(brief, package, resources, market_snapshot)
    wb = load_workbook(BytesIO(base))
    preview_count = add_preview_sheet(wb, package)
    if "Briefing" in wb.sheetnames:
        ws = wb["Briefing"]
        ws.append(["mode_exporter", MODE_EXPORTER_VERSION])
        ws.append(["embedded_previews", preview_count])
    return _save_verified(wb, preview_count)


def _story_context_rows(package: dict) -> list[dict]:
    context = package.get("story_context") or {}
    hero = context.get("hero_story") or {}
    return [
        {"item": "pipeline", "value": (package.get("content_quality") or {}).get("pipeline")},
        {"item": "hero_story", "value": hero.get("headline_ja")},
        {"item": "archetype", "value": hero.get("archetype")},
        {"item": "story_score", "value": hero.get("story_score")},
        {"item": "why_now", "value": hero.get("why_now_ja")},
        {"item": "conflict", "value": hero.get("conflict_ja")},
        {"item": "implication", "value": hero.get("implication_ja")},
        {"item": "generation_seed", "value": (package.get("content_quality") or {}).get("generation_seed")},
    ]


def _story_candidate_rows(package: dict) -> list[dict]:
    rows: list[dict] = []
    for candidate in (package.get("story_context") or {}).get("candidates") or []:
        rows.append(
            {
                "id": candidate.get("id"),
                "headline_ja": candidate.get("headline_ja"),
                "archetype": candidate.get("archetype"),
                "topic": candidate.get("topic"),
                "story_score": candidate.get("story_score"),
                "confidence": candidate.get("confidence"),
                "cluster_size": candidate.get("cluster_size"),
                "sources": candidate.get("source_names"),
                "why_now_ja": candidate.get("why_now_ja"),
                "conflict_ja": candidate.get("conflict_ja"),
                "implication_ja": candidate.get("implication_ja"),
            }
        )
    return rows


def _story_card_rows(package: dict) -> list[dict]:
    rows: list[dict] = []
    for set_name, card in _renderable_cards(package):
        rows.append(
            {
                "set": set_name,
                "slide": card.get("slide"),
                "story_role": card.get("story_role"),
                "story_archetype": card.get("story_archetype"),
                "headline": card.get("headline"),
                "body": card.get("key_message"),
                "evidence_refs": card.get("evidence_refs"),
                "source": card.get("source"),
                "layout": (card.get("visual_direction") or {}).get("layout_variant"),
                "visual_prompt_4_5": ((card.get("visual_direction") or {}).get("image_prompts") or {}).get("4:5"),
            }
        )
    return rows


def build_story_excel(package: dict, resources: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Story_Context"
    _append_rows(ws, ["item", "value"], _story_context_rows(package))

    candidates = wb.create_sheet("Story_Candidates")
    _append_rows(
        candidates,
        ["id", "headline_ja", "archetype", "topic", "story_score", "confidence", "cluster_size", "sources", "why_now_ja", "conflict_ja", "implication_ja"],
        _story_candidate_rows(package),
    )

    cards = wb.create_sheet("Story_Cards")
    _append_rows(
        cards,
        ["set", "slide", "story_role", "story_archetype", "headline", "body", "evidence_refs", "source", "layout", "visual_prompt_4_5"],
        _story_card_rows(package),
    )

    sources = wb.create_sheet("Sources")
    _append_rows(
        sources,
        ["id", "source", "source_type", "region", "category", "title", "story_score", "trader_score", "risk_score", "tags", "url", "excerpt"],
        resources,
    )

    note = wb.create_sheet("Note")
    note.append(["markdown"])
    note.append([package.get("note_markdown") or ""])
    _style_table(note)

    preview_count = add_preview_sheet(wb, package)
    wb["Story_Context"].append(["embedded_previews", preview_count])
    return _save_verified(wb, preview_count)
