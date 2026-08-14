from __future__ import annotations

import json
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from market_data import flatten_derivatives_rows, flatten_indicator_rows, flatten_level_rows, flatten_market_rows


HEADER_FILL = PatternFill("solid", fgColor="E8EEF7")
TITLE_FILL = PatternFill("solid", fgColor="102A43")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=13)
HEADER_FONT = Font(bold=True, color="243B53")


def stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def style_sheet(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.fill = TITLE_FILL
        cell.font = TITLE_FONT
    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells[:80]:
            max_len = max(max_len, len(stringify(cell.value)))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 44)


def append_table(ws, headers: list[str], rows: list[dict]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in rows:
        ws.append([stringify(row.get(header, "")) for header in headers])
    style_sheet(ws)


def add_brief_sheet(wb: Workbook, brief: dict) -> None:
    ws = wb.active
    ws.title = "Briefing"
    ws.append(["항목", "내용"])
    rows = [
        ("제목", brief.get("title", "")),
        ("한 줄 판단", brief.get("one_line", "")),
        ("생성 방식", brief.get("provider", "")),
        ("시장 요약", brief.get("market_summary", {})),
        ("리스크", "\n".join(brief.get("risk_notes", []))),
    ]
    for key, value in rows:
        ws.append([key, stringify(value)])
    ws.append(["핵심 포인트", "\n".join(brief.get("key_points", []))])
    ws.append(["트레이더 문장", "\n".join(brief.get("trader_sentences", []))])
    ws.append(["시장 구조", stringify(brief.get("market_structure", {}))])
    ws.append(["무효화 조건", "\n".join(brief.get("invalidation_points", []))])
    ws.append(["실행 체크리스트", "\n".join(brief.get("action_plan", []))])
    style_sheet(ws)


def add_source_findings_sheet(wb: Workbook, brief: dict) -> None:
    ws = wb.create_sheet("Source_Findings")
    findings = brief.get("source_findings", []) or []
    rows = []
    for item in findings:
        rows.append(
            {
                "source": item.get("source", ""),
                "title": item.get("title", ""),
                "role": item.get("role", ""),
                "material_chars": item.get("material_chars", ""),
                "evidence": "\n".join(item.get("evidence", [])),
                "trader_read": item.get("trader_read", ""),
                "url": item.get("url", ""),
            }
        )
    append_table(ws, ["source", "title", "role", "material_chars", "evidence", "trader_read", "url"], rows)


def add_scenarios_sheet(wb: Workbook, brief: dict) -> None:
    ws = wb.create_sheet("Scenarios")
    append_table(
        ws,
        ["case", "probability_view", "trigger", "expected_path", "watch"],
        brief.get("scenarios", []) or [],
    )


def add_card_sheet(wb: Workbook, name: str, cards: list[dict]) -> None:
    ws = wb.create_sheet(name[:31])
    append_table(
        ws,
        ["set", "slide", "headline", "body", "caption", "visual_direction", "source_hint"],
        cards,
    )


def add_note_sheet(wb: Workbook, note_markdown: str) -> None:
    ws = wb.create_sheet("Note")
    ws.append(["markdown"])
    ws.append([note_markdown])
    style_sheet(ws)


def add_sources_sheet(wb: Workbook, resources: list[dict]) -> None:
    ws = wb.create_sheet("Sources")
    append_table(
        ws,
        ["source", "source_type", "region", "category", "title", "tags", "trader_score", "risk_score", "url", "excerpt"],
        resources,
    )


def add_market_sheet(wb: Workbook, market_snapshot: dict) -> None:
    ws = wb.create_sheet("Market")
    append_table(
        ws,
        [
            "name",
            "symbol",
            "asset_class",
            "price",
            "unit",
            "change_24h",
            "change_7d",
            "change_30d",
            "technical_bias",
            "nearest_support",
            "nearest_resistance",
            "rsi14",
            "macd_bias",
            "market_cap",
            "source",
        ],
        flatten_market_rows(market_snapshot),
    )


def add_price_levels_sheet(wb: Workbook, market_snapshot: dict) -> None:
    ws = wb.create_sheet("Price_Levels")
    append_table(
        ws,
        ["asset", "direction", "level", "distance_pct", "reason", "importance", "source"],
        flatten_level_rows(market_snapshot),
    )


def add_indicators_sheet(wb: Workbook, market_snapshot: dict) -> None:
    ws = wb.create_sheet("Indicators")
    append_table(
        ws,
        [
            "asset",
            "current",
            "ma20",
            "ma50",
            "ma100",
            "ma200",
            "ema20",
            "rsi14",
            "macd",
            "macd_signal",
            "macd_histogram",
            "macd_bias",
            "bollinger_upper",
            "bollinger_middle",
            "bollinger_lower",
            "bollinger_bandwidth_pct",
            "atr14",
            "atr14_pct",
            "volume_20d_avg",
        ],
        flatten_indicator_rows(market_snapshot),
    )


def add_derivatives_sheet(wb: Workbook, market_snapshot: dict) -> None:
    ws = wb.create_sheet("Derivatives")
    append_table(
        ws,
        ["pair", "mark_price", "index_price", "last_funding_rate", "next_funding_time", "open_interest_contracts", "source", "error"],
        flatten_derivatives_rows(market_snapshot),
    )


def build_excel_bytes(brief: dict, content_package: dict, resources: list[dict], market_snapshot: dict) -> bytes:
    wb = Workbook()
    add_brief_sheet(wb, brief)
    add_source_findings_sheet(wb, brief)
    add_scenarios_sheet(wb, brief)
    for name, cards in (content_package.get("cards") or {}).items():
        add_card_sheet(wb, f"Cards_{name}", cards)
    add_note_sheet(wb, content_package.get("note_markdown", ""))
    add_sources_sheet(wb, resources)
    add_market_sheet(wb, market_snapshot)
    add_price_levels_sheet(wb, market_snapshot)
    add_indicators_sheet(wb, market_snapshot)
    add_derivatives_sheet(wb, market_snapshot)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
