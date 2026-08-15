from __future__ import annotations

import copy
import html as html_lib
import re
from io import BytesIO
from math import ceil
from pathlib import Path
from typing import Any

from card_renderer import render_card_png

_PREVIEW_RE = re.compile(r'<div class="observer-preview\s+([^\"]+)">', re.S)
_FIELD_RE = {
    "eyebrow": re.compile(r'<div class="observer-eyebrow">(.*?)</div>', re.S),
    "headline": re.compile(r'<div class="observer-headline">(.*?)</div>', re.S),
    "subheadline": re.compile(r'<div class="observer-sub">(.*?)</div>', re.S),
    "key_message": re.compile(r'<div class="observer-message">(.*?)</div>', re.S),
    "footer": re.compile(r'<div class="observer-source">(.*?)</div>', re.S),
}
_METRIC_RE = re.compile(r"<div class='observer-metric\s*([^']*)'><span>(.*?)</span><strong>(.*?)</strong></div>", re.S)


def _clean_html_text(value: str) -> str:
    value = re.sub(r"<[^>]+>", "", value or "")
    return html_lib.unescape(value).strip()


def _parse_preview_html(markup: str) -> dict | None:
    layout_match = _PREVIEW_RE.search(markup or "")
    if not layout_match:
        return None
    parsed: dict[str, Any] = {"layout": _clean_html_text(layout_match.group(1)), "metrics": []}
    for key, pattern in _FIELD_RE.items():
        match = pattern.search(markup)
        parsed[key] = _clean_html_text(match.group(1)) if match else ""
    for variant, label, value in _METRIC_RE.findall(markup):
        parsed["metrics"].append({"id": "", "label": _clean_html_text(label), "value": _clean_html_text(value), "variant": _clean_html_text(variant)})
    return parsed


def _all_session_cards(st) -> list[dict]:
    try:
        package = st.session_state.get("content_package") or {}
    except Exception:
        return []
    cards: list[dict] = []
    for card_set in (package.get("cards") or {}).values():
        cards.extend(card_set or [])
    return cards


def _match_card(st, parsed: dict) -> dict:
    headline = parsed.get("headline", "")
    footer = parsed.get("footer", "")
    candidates = [card for card in _all_session_cards(st) if str(card.get("headline") or "") == headline]
    if footer and len(candidates) > 1:
        for card in candidates:
            if str(card.get("footer") or "") == footer:
                return card
    if candidates:
        return candidates[0]
    layout = str(parsed.get("layout") or "")
    metric_labels = " ".join(str(m.get("label") or "").upper() for m in parsed.get("metrics") or [])
    footer_lower = footer.lower()
    if "brand_outro" in layout or "follow" in footer_lower:
        card_type = "brand_outro"
    elif "news_primary" in layout or any(name in footer_lower for name in ["coindesk", "cointelegraph", "u.today", "decrypt", "yahoo"]):
        card_type = "news_context"
    elif "FUNDING" in metric_labels or "OI" in metric_labels or "RSI" in metric_labels:
        card_type = "derivatives"
    elif "scenario_primary" in layout or "MA20" in metric_labels or "MA50" in metric_labels:
        card_type = "scenarios"
    elif "chart_primary" in layout or "SUPPORT" in metric_labels or "RESISTANCE" in metric_labels:
        card_type = "key_levels"
    elif any(token in headline for token in ["条件", "入らない", "固定", "間違えたら"]):
        card_type = "trade_plan"
    else:
        card_type = "market_conclusion"
    return {
        "card_type": card_type,
        "eyebrow": parsed.get("eyebrow") or "キヨサキ",
        "headline": headline,
        "subheadline": parsed.get("subheadline"),
        "key_message": parsed.get("key_message"),
        "metrics": parsed.get("metrics") or [],
        "footer": footer,
        "source": {},
    }


def _metric_value(card: dict, *metric_ids: str) -> str:
    by_id = {str(item.get("id")): item for item in (card.get("metrics") or []) if item.get("id")}
    for metric_id in metric_ids:
        value = str((by_id.get(metric_id) or {}).get("value") or "").strip()
        if value:
            return value
    return ""


def _news_headline(card: dict) -> str:
    source = card.get("source") or {}
    localized = str(source.get("display_headline_ja") or "").strip()
    if localized:
        return localized
    raw = str(source.get("short_title") or "").strip()
    low = raw.lower()
    if "supply" in low and ("bitcoin" in low or "btc" in low):
        return "BTC供給の希少性が再び焦点に"
    if "etf" in low and ("flow" in low or "inflow" in low):
        return "ETF資金フローに再び注目"
    if "chainlink" in low:
        return "Chainlinkへの期待が再浮上"
    if "regulation" in low or "sec" in low:
        return "規制材料が市場の焦点に"
    return str(card.get("headline") or raw or "市場材料を確認")


def _normalize_trade_plan(card: dict) -> None:
    resistance = _metric_value(card, "btc_primary_resistance", "btc_resistance")
    support = _metric_value(card, "btc_primary_support", "btc_support")
    plan = copy.deepcopy(card.get("trade_plan") or {})
    plan["entry"] = {"visible": True, "condition": f"{resistance}を終値で回復" if resistance else "上値条件を終値で回復"}
    plan["wait"] = {"visible": True, "condition": "レンジ内は待機"}
    plan["invalid"] = {"visible": True, "condition": f"{support}を終値で割る" if support else "無効化条件を終値で割る"}
    card["trade_plan"] = plan
    card["headline"] = "条件を先に固定する"
    card["key_message"] = "入る条件より、入らない条件を先に決める。"


def _normalize_scenario(card: dict) -> None:
    ma50 = _metric_value(card, "ma50")
    ma20 = _metric_value(card, "ma20")
    resistance = _metric_value(card, "btc_primary_resistance", "btc_resistance")
    hierarchy = [value for value in [ma50, ma20, resistance] if value]
    if hierarchy:
        card["key_message"] = f"上は{' → '.join(hierarchy)}。近い順に回復を確認する。"
    card["headline"] = "突破は点ではなく帯で見る"


def _normalize_news(card: dict) -> None:
    card["headline"] = _news_headline(card)
    reaction = ((card.get("source") or {}).get("news_reaction") or {})
    if reaction.get("available"):
        card["key_message"] = str(card.get("key_message") or "材料だけでなく、実際の価格反応を見る。")
    else:
        card["key_message"] = "反応データがないため、材料と現在位置を分けて見る。"


def _dedupe_and_editorialize_cards(cards: list[dict]) -> list[dict]:
    source = [copy.deepcopy(card) for card in (cards or [])]
    trade_indexes = [i for i, card in enumerate(source) if card.get("card_type") == "trade_plan"]
    keep_trade_index = trade_indexes[-1] if trade_indexes else None
    cleaned: list[dict] = []
    for index, card in enumerate(source):
        if card.get("card_type") == "trade_plan" and index != keep_trade_index:
            continue
        card_type = card.get("card_type")
        if card_type == "trade_plan":
            _normalize_trade_plan(card)
        elif card_type == "scenarios":
            _normalize_scenario(card)
        elif card_type == "news_context":
            _normalize_news(card)
        cleaned.append(card)
    for slide, card in enumerate(cleaned, start=1):
        card["slide"] = slide
    return cleaned


def _postprocess_content_package(package: dict) -> dict:
    next_package = copy.deepcopy(package or {})
    cards_by_set = next_package.get("cards") or {}
    for set_label, cards in list(cards_by_set.items()):
        cards_by_set[set_label] = _dedupe_and_editorialize_cards(cards)
    quality = next_package.get("content_quality") or {}
    quality["preview_renderer"] = "documentary-editorial-v4"
    quality["visual_asset_policy"] = "full-bleed generated/real asset when available; cinematic deterministic fallback otherwise"
    quality["trade_plan_policy"] = "one execution card maximum per carousel"
    next_package["content_quality"] = quality
    return next_package


def _install_reasoning_package_patch() -> None:
    import reasoning_engine
    if getattr(reasoning_engine, "_kiyosaki_package_postprocess_applied", False):
        return
    original = reasoning_engine.generate_content_package

    def generate_content_package(*args, **kwargs):
        return _postprocess_content_package(original(*args, **kwargs))

    reasoning_engine.generate_content_package = generate_content_package
    reasoning_engine._kiyosaki_package_postprocess_applied = True


def _install_markdown_preview_patch(st) -> None:
    if getattr(st, "_kiyosaki_preview_renderer_applied", False):
        return
    original_markdown = st.markdown
    original_image = st.image

    def patched_markdown(body, *args, **kwargs):
        markup = str(body or "")
        parsed = _parse_preview_html(markup)
        if parsed:
            card = _match_card(st, parsed)
            try:
                png = render_card_png(card, width=1080, height=1350)
                return original_image(png, width=430)
            except Exception:
                safe_markup = re.sub(r"<img\b[^>]*>", "", markup, flags=re.I | re.S)
                safe_markup = safe_markup.replace("observer-figure", "observer-figure preview-image-disabled")
                return original_markdown(safe_markup, *args, **kwargs)
        return original_markdown(body, *args, **kwargs)

    st.markdown = patched_markdown
    st._kiyosaki_preview_renderer_applied = True


def _disable_reference_asset_reads() -> None:
    if getattr(Path, "_kiyosaki_reference_read_patch", False):
        return
    original_exists = Path.exists

    def patched_exists(self: Path) -> bool:
        if self.name == "observer_reference.png" and "assets" in self.parts and "brand" in self.parts:
            return False
        return original_exists(self)

    Path.exists = patched_exists
    Path._kiyosaki_reference_read_patch = True


def _add_preview_sheet(workbook, content_package: dict) -> None:
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill

    if "Card_Previews" in workbook.sheetnames:
        del workbook["Card_Previews"]
    ws = workbook.create_sheet("Card_Previews")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    for column_letter in "ABCDEFGHIJKLMN":
        ws.column_dimensions[column_letter].width = 11
    title_fill = PatternFill("solid", fgColor="111111")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    label_font = Font(color="E68A19", bold=True, size=11)
    note_font = Font(color="666666", italic=True, size=9)
    ws.merge_cells("A1:N1")
    ws["A1"] = "Kiyosaki Card Preview Gallery"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    row_cursor = 3
    for set_label, cards in (content_package.get("cards") or {}).items():
        production_cards = [card for card in (cards or []) if (card.get("qa") or {}).get("renderable", True)]
        if not production_cards:
            continue
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=14)
        heading = ws.cell(row=row_cursor, column=1)
        heading.value = f"{set_label} · {len(production_cards)} cards"
        heading.font = label_font
        row_cursor += 2
        for index, card in enumerate(production_cards):
            grid_col = 1 if index % 2 == 0 else 8
            grid_row = row_cursor + (index // 2) * 22
            label = ws.cell(row=grid_row, column=grid_col)
            label.value = f"{card.get('slide', index + 1)}. {card.get('headline', '')}"
            label.font = Font(bold=True, size=10)
            label.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=grid_row, start_column=grid_col, end_row=grid_row, end_column=grid_col + 5)
            png = render_card_png(card, width=540, height=675)
            image = XLImage(BytesIO(png))
            image.width = 270
            image.height = 338
            ws.add_image(image, ws.cell(row=grid_row + 1, column=grid_col).coordinate)
        row_cursor += ceil(len(production_cards) / 2) * 22 + 2
    ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=14)
    ws.cell(row=row_cursor, column=1).value = "Preview images use the same deterministic renderer as Streamlit and contain no Observer reference-sheet image."
    ws.cell(row=row_cursor, column=1).font = note_font
    ws.cell(row=row_cursor, column=1).alignment = Alignment(wrap_text=True)


def _install_excel_preview_patch(excel_exporter) -> None:
    if getattr(excel_exporter, "_kiyosaki_excel_preview_applied", False):
        return
    original_build = excel_exporter.build_excel_bytes

    def build_excel_bytes(brief: dict, content_package: dict, resources: list[dict], market_snapshot: dict) -> bytes:
        from openpyxl import load_workbook
        base = original_build(brief, content_package, resources, market_snapshot)
        try:
            workbook = load_workbook(BytesIO(base))
            _add_preview_sheet(workbook, content_package)
            output = BytesIO()
            workbook.save(output)
            return output.getvalue()
        except Exception:
            return base

    excel_exporter.build_excel_bytes = build_excel_bytes
    excel_exporter._kiyosaki_excel_preview_applied = True


def apply_preview_runtime() -> None:
    import streamlit as st
    import excel_exporter
    _install_reasoning_package_patch()
    _disable_reference_asset_reads()
    _install_markdown_preview_patch(st)
    _install_excel_preview_patch(excel_exporter)
