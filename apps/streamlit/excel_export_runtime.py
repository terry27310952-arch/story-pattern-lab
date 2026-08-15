from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from excel_preview_sheet import PREVIEW_SHEET_NAME, add_card_preview_sheet


EXCEL_PREVIEW_RUNTIME_VERSION = "excel-preview-v4.2"


def apply_excel_export_patch() -> None:
    """Install a versioned Excel exporter wrapper before app.py imports build_excel_bytes.

    This is intentionally separate from preview_runtime because app.py uses
    `from excel_exporter import build_excel_bytes`. The wrapper must therefore
    be installed on the excel_exporter module before app.py is executed.
    """
    import excel_exporter

    current_version = getattr(excel_exporter, "_kiyosaki_excel_export_runtime_version", None)
    if current_version == EXCEL_PREVIEW_RUNTIME_VERSION:
        return

    base_builder = getattr(excel_exporter, "_kiyosaki_excel_export_base_builder", None)
    if base_builder is None:
        base_builder = excel_exporter.build_excel_bytes
        excel_exporter._kiyosaki_excel_export_base_builder = base_builder

    def build_excel_bytes(brief: dict, content_package: dict, resources: list[dict], market_snapshot: dict) -> bytes:
        base = base_builder(brief, content_package, resources, market_snapshot)
        workbook = load_workbook(BytesIO(base))
        embedded = add_card_preview_sheet(workbook, content_package)

        # Make the feature auditable inside the workbook itself.
        if "Briefing" in workbook.sheetnames:
            ws = workbook["Briefing"]
            row = ws.max_row + 2
            ws.cell(row=row, column=1).value = "Card preview export"
            ws.cell(row=row, column=2).value = f"{EXCEL_PREVIEW_RUNTIME_VERSION} / embedded images: {embedded}"

        output = BytesIO()
        workbook.save(output)
        payload = output.getvalue()

        # Fail loudly instead of silently returning a workbook without previews.
        verify = load_workbook(BytesIO(payload))
        if PREVIEW_SHEET_NAME not in verify.sheetnames:
            raise RuntimeError("Card_Previews sheet was not created")
        if embedded > 0 and len(verify[PREVIEW_SHEET_NAME]._images) != embedded:
            raise RuntimeError(
                f"Card preview image verification failed: expected {embedded}, "
                f"found {len(verify[PREVIEW_SHEET_NAME]._images)}"
            )
        return payload

    excel_exporter.build_excel_bytes = build_excel_bytes
    excel_exporter._kiyosaki_excel_export_runtime_version = EXCEL_PREVIEW_RUNTIME_VERSION
