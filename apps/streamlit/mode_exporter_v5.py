from __future__ import annotations

import json
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font

import card_renderer
import mode_exporter_v4 as legacy
import story_renderer_v5


MODE_EXPORTER_VERSION = "mode-exporter-v10.0"
build_trader_excel = legacy.build_trader_excel


def _renderable_cards(package: dict) -> list[tuple[str, dict]]:
    out: list[tuple[str, dict]] = []
    for set_name, cards in (package.get("cards") or {}).items():
        for card in cards or []:
            if (card.get("qa") or {}).get("renderable", True):
                out.append((str(set_name), card))
    return out


def render_card_by_mode(package: dict, card: dict, width: int, height: int) -> bytes:
    if package.get("mode") == "story" or (card.get("qa") or {}).get("mode") == "story":
        return story_renderer_v5.render_story_card_png(card, width=width, height=height)
    return card_renderer.render_card_png(card, width=width, height=height)


def add_preview_sheet(wb: Workbook, package: dict) -> int:
    if "Card_Previews" in wb.sheetnames:
        del wb["Card_Previews"]
    ws = wb.create_sheet("Card_Previews", 0)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "キヨサキ · Card Previews"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = "このシートの画像は、アプリと同一のStory v10レンダラーで生成されています。"
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 44

    streams: list[BytesIO] = []
    count = 0
    row_cursor = 4
    for index, (set_name, card) in enumerate(_renderable_cards(package)):
        col = "A" if index % 2 == 0 else "B"
        if index and index % 2 == 0:
            row_cursor += 22
        ws[f"{col}{row_cursor}"] = f"{set_name} · {card.get('slide')} · {card.get('headline', '')}"
        ws[f"{col}{row_cursor}"].font = Font(bold=True)
        ws[f"{col}{row_cursor + 1}"] = f"{card.get('story_role') or card.get('card_type')} · {(card.get('visual_direction') or {}).get('scene_type','')}"
        stream = BytesIO(render_card_by_mode(package, card, 540, 675))
        streams.append(stream)
        image = XLImage(stream)
        image.width = 270
        image.height = 337
        ws.add_image(image, f"{col}{row_cursor + 2}")
        count += 1
    wb._kiyosaki_preview_streams = streams  # type: ignore[attr-defined]
    return count


def _context_rows(package: dict) -> list[dict]:
    context = package.get("story_context") or {}
    hero = context.get("hero_story") or {}
    quality = package.get("content_quality") or {}
    plan = context.get("story_plan") or {}
    return [
        {"item": "pipeline", "value": quality.get("pipeline")},
        {"item": "source_engine", "value": quality.get("source_engine")},
        {"item": "graph_engine", "value": quality.get("graph_engine")},
        {"item": "renderer", "value": quality.get("renderer")},
        {"item": "hero_story", "value": plan.get("headline_ja") or hero.get("headline_ja")},
        {"item": "archetype_tag", "value": plan.get("archetype_tag")},
        {"item": "story_score", "value": hero.get("story_score")},
        {"item": "hero_story_score", "value": hero.get("hero_story_score")},
        {"item": "hero_resource_ids", "value": context.get("hero_resource_ids")},
        {"item": "cluster_event_scores", "value": context.get("cluster_event_scores")},
        {"item": "story_thesis", "value": plan.get("thesis")},
        {"item": "planning_policy", "value": plan.get("planning_policy")},
        {"item": "story_qa", "value": quality.get("story_qa")},
        {"item": "generation_seed", "value": quality.get("generation_seed")},
    ]


def _candidate_rows(package: dict) -> list[dict]:
    rows = []
    for c in (package.get("story_context") or {}).get("candidates") or []:
        rows.append({
            "id": c.get("id"), "headline_seed": c.get("headline_seed"), "topic": c.get("topic"),
            "source_archetype_hint": c.get("archetype"), "story_score": c.get("story_score"),
            "hero_story_score": c.get("hero_story_score"), "confidence": c.get("confidence"),
            "cluster_size": c.get("cluster_size"), "cluster_coherence": c.get("cluster_coherence"),
            "event_fingerprint": c.get("event_fingerprint"), "entities": c.get("entities"), "sources": c.get("source_names"),
        })
    return rows


def _graph_rows(package: dict) -> list[dict]:
    graph = (package.get("story_context") or {}).get("fact_graph") or {}
    return [
        {
            "id": f.get("id"), "source_id": f.get("source_id"), "index": f.get("index"),
            "subject": f.get("subject"), "relation": f.get("relation"), "values": f.get("values"),
            "years": f.get("years"), "score": f.get("score"), "sentence": f.get("sentence"),
        }
        for f in graph.get("facts") or []
    ]


def _plan_rows(package: dict) -> list[dict]:
    plan = (package.get("story_context") or {}).get("story_plan") or {}
    return [
        {"order": idx + 1, "role": item.get("role"), "fact_ids": item.get("fact_ids"), "scene_type": item.get("scene_type"), "reason": item.get("reason")}
        for idx, item in enumerate(plan.get("cards") or [])
    ]


def _card_rows(package: dict) -> list[dict]:
    rows = []
    for set_name, card in _renderable_cards(package):
        direction = card.get("visual_direction") or {}
        rows.append({
            "set": set_name, "slide": card.get("slide"), "story_role": card.get("story_role"),
            "story_archetype": card.get("story_archetype"), "headline": card.get("headline"),
            "body": card.get("key_message"), "evidence_excerpt": card.get("evidence_excerpt"),
            "evidence_refs": card.get("evidence_refs"), "fact_bound": (card.get("qa") or {}).get("fact_bound"),
            "claim_evidence_consistent": (card.get("qa") or {}).get("claim_evidence_consistent"),
            "layout": direction.get("layout_variant"), "scene_type": direction.get("scene_type"),
            "visual_prompt_4_5": (direction.get("image_prompts") or {}).get("4:5"), "source": card.get("source"),
        })
    return rows


def build_story_excel(package: dict, resources: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Story_Context"
    legacy._append_rows(ws, ["item", "value"], _context_rows(package))

    candidates = wb.create_sheet("Story_Candidates")
    legacy._append_rows(candidates, ["id", "headline_seed", "topic", "source_archetype_hint", "story_score", "hero_story_score", "confidence", "cluster_size", "cluster_coherence", "event_fingerprint", "entities", "sources"], _candidate_rows(package))

    graph = wb.create_sheet("Story_Graph")
    legacy._append_rows(graph, ["id", "source_id", "index", "subject", "relation", "values", "years", "score", "sentence"], _graph_rows(package))

    plan = wb.create_sheet("Story_Plan")
    legacy._append_rows(plan, ["order", "role", "fact_ids", "scene_type", "reason"], _plan_rows(package))

    cards = wb.create_sheet("Story_Cards")
    legacy._append_rows(cards, ["set", "slide", "story_role", "story_archetype", "headline", "body", "evidence_excerpt", "evidence_refs", "fact_bound", "claim_evidence_consistent", "layout", "scene_type", "visual_prompt_4_5", "source"], _card_rows(package))

    qa = wb.create_sheet("Story_QA")
    story_qa = (package.get("content_quality") or {}).get("story_qa") or {}
    legacy._append_rows(qa, ["item", "value"], [{"item": key, "value": value} for key, value in story_qa.items()])

    sources = wb.create_sheet("Sources")
    legacy._append_rows(sources, ["id", "source", "source_type", "region", "category", "title", "story_score", "trader_score", "risk_score", "tags", "url", "excerpt"], resources)

    note = wb.create_sheet("Note")
    note.append(["markdown"])
    note.append([package.get("note_markdown") or ""])
    legacy._style_table(note)

    preview_count = add_preview_sheet(wb, package)
    wb["Story_Context"].append(["embedded_previews", preview_count])
    return legacy._save_verified(wb, preview_count)
