from __future__ import annotations

import json
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


STORY_EXPORT_RUNTIME_VERSION = "story-export-v5.0"
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ACCENT_FILL = PatternFill("solid", fgColor="F59E0B")


def _string(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _style(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        width = 12
        for cell in list(column_cells)[:80]:
            width = max(width, min(60, len(_string(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _add_story_context(wb, brief: dict, content_package: dict) -> None:
    if "Story_Context" in wb.sheetnames:
        del wb["Story_Context"]
    ws = wb.create_sheet("Story_Context", 1)
    ws.append(["item", "value"])
    story_context = brief.get("story_context") or {}
    hero = story_context.get("hero_story") or {}
    quality = content_package.get("content_quality") or {}
    story_meta = quality.get("story_engine") or {}
    rows = [
        ("runtime", story_meta.get("runtime", STORY_EXPORT_RUNTIME_VERSION)),
        ("hero_story", hero.get("headline_ja") or quality.get("hero_story_title") or ""),
        ("archetype", hero.get("archetype") or quality.get("story_archetype") or ""),
        ("story_score", hero.get("story_score") or quality.get("story_score") or 0),
        ("confidence", hero.get("confidence") or ""),
        ("topic", hero.get("topic") or ""),
        ("entities", hero.get("entities") or []),
        ("why_now", hero.get("why_now_ja") or ""),
        ("conflict", hero.get("conflict_ja") or ""),
        ("market_implication", hero.get("implication_ja") or ""),
        ("visual_motifs", hero.get("visual_motifs") or []),
        ("source", hero.get("hero_resource") or {}),
        ("policy", story_context.get("policy") or story_meta.get("policy") or ""),
    ]
    for key, value in rows:
        ws.append([key, _string(value)])
    _style(ws)


def _add_story_candidates(wb, brief: dict) -> None:
    if "Story_Candidates" in wb.sheetnames:
        del wb["Story_Candidates"]
    ws = wb.create_sheet("Story_Candidates", 2)
    headers = ["rank", "story_score", "archetype", "headline_ja", "topic", "entities", "cluster_size", "confidence", "source_names", "headline_seed"]
    ws.append(headers)
    candidates = (brief.get("story_context") or {}).get("candidates") or []
    for index, item in enumerate(candidates, start=1):
        ws.append([
            index,
            item.get("story_score", ""),
            item.get("archetype", ""),
            item.get("headline_ja", ""),
            item.get("topic", ""),
            _string(item.get("entities") or []),
            item.get("cluster_size", ""),
            item.get("confidence", ""),
            _string(item.get("source_names") or []),
            item.get("headline_seed", ""),
        ])
    _style(ws)


def _add_story_cards(wb, content_package: dict) -> None:
    if "Story_Cards" in wb.sheetnames:
        del wb["Story_Cards"]
    ws = wb.create_sheet("Story_Cards", 3)
    headers = ["set", "slide", "story_archetype", "story_role", "card_type", "headline", "subheadline", "layout", "visual_focus", "source"]
    ws.append(headers)
    for set_label, cards in (content_package.get("cards") or {}).items():
        for card in cards or []:
            direction = card.get("visual_direction") or {}
            source = card.get("source") or {}
            ws.append([
                set_label,
                card.get("slide", ""),
                card.get("story_archetype", ""),
                card.get("story_role", ""),
                card.get("card_type", ""),
                card.get("headline", ""),
                card.get("subheadline", ""),
                direction.get("format_variant") or direction.get("layout_variant") or "",
                direction.get("visual_focus", ""),
                " · ".join(value for value in [source.get("publisher", ""), source.get("short_title", "")] if value),
            ])
    _style(ws)


def apply_excel_patch(excel_exporter) -> None:
    if getattr(excel_exporter, "_kiyosaki_story_export_version", None) == STORY_EXPORT_RUNTIME_VERSION:
        return
    original = excel_exporter.build_excel_bytes

    def build_excel_bytes(brief: dict, content_package: dict, resources: list[dict], market_snapshot: dict) -> bytes:
        raw = original(brief, content_package, resources, market_snapshot)
        wb = load_workbook(BytesIO(raw))
        _add_story_context(wb, brief, content_package)
        _add_story_candidates(wb, brief)
        _add_story_cards(wb, content_package)
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    excel_exporter.build_excel_bytes = build_excel_bytes
    excel_exporter._kiyosaki_story_export_version = STORY_EXPORT_RUNTIME_VERSION
